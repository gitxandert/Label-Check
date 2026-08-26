import csv
import datetime
import tempfile
import unittest
from unittest import mock
from pathlib import Path
import sys


SRC_DIR = Path(__file__).resolve().parents[1] / "src"
sys.path.insert(0, str(SRC_DIR))

import app as app_module
from openpyxl import Workbook, load_workbook


class StatisticsStoreTests(unittest.TestCase):
    def setUp(self):
        self.temporary = tempfile.TemporaryDirectory()
        self.root = Path(self.temporary.name)
        self.store = app_module.StatisticsStore(
            str(self.root / "statistics.sqlite3"), str(self.root / "users")
        )

    def tearDown(self):
        self.temporary.cleanup()

    def test_rollup_only_writes_days_with_authenticated_presence(self):
        present = datetime.date(2026, 8, 20)
        absent = datetime.date(2026, 8, 21)
        active = datetime.date(2026, 8, 22)
        self.store.note_presence("operator", present)
        self.store.increment("operator", "slides_completed", 2, active)

        path = self.store.rollup_user("operator", active)

        with open(path, newline="", encoding="utf-8") as handle:
            rows = list(csv.DictReader(handle))
        self.assertEqual([row["date"] for row in rows], [present.isoformat(), active.isoformat()])
        self.assertNotIn(absent.isoformat(), [row["date"] for row in rows])
        self.assertEqual(rows[0]["slides_completed"], "0")
        self.assertEqual(rows[1]["slides_completed"], "2")

    def test_hours_round_half_up(self):
        self.assertEqual(self.store.rounded_hours(29), 0)
        self.assertEqual(self.store.rounded_hours(30), 1)
        self.assertEqual(self.store.rounded_hours(89), 1)
        self.assertEqual(self.store.rounded_hours(90), 2)

    def test_heartbeat_credits_at_most_once_per_minute(self):
        first = datetime.datetime(2026, 8, 20, 9, 0, 5)
        self.assertTrue(self.store.heartbeat("operator", first))
        self.assertFalse(
            self.store.heartbeat("operator", first + datetime.timedelta(seconds=40))
        )
        self.assertTrue(
            self.store.heartbeat("operator", first + datetime.timedelta(minutes=1))
        )
        row = self.store._database_rows("operator")["2026-08-20"]
        self.assertEqual(row["active_minutes"], 2)

    def test_dashboard_merges_live_today_and_fills_seven_days(self):
        today = datetime.date(2026, 8, 25)
        self.store.local_date = lambda: today
        self.store.increment("operator", "accessions_logged", 3, today)
        for minute in range(30):
            self.store.heartbeat(
                "operator", datetime.datetime(2026, 8, 25, 9, minute)
            )

        dashboard = self.store.dashboard("operator")

        self.assertEqual(len(dashboard["week"]), 7)
        self.assertEqual(dashboard["week"][0]["date"], "2026-08-19")
        self.assertEqual(dashboard["week"][-1]["accessions_logged"], 3)
        self.assertEqual(dashboard["week"][-1]["hours"], 1)
        self.assertEqual(dashboard["totals"]["accessions_logged"], 3)

    def test_storage_key_cannot_escape_user_root(self):
        path = self.store.csv_path("../../another-user")
        self.assertEqual(path.parent.name, "logs")
        self.assertEqual(path.parents[2], self.root / "users")


class SDLStatisticsIntegrationTests(unittest.TestCase):
    def setUp(self):
        self.temporary = tempfile.TemporaryDirectory()
        self.root = Path(self.temporary.name)
        self.old_users = app_module.user_manager.users.copy()
        self.old_sdl_path = app_module.Config.SDL_FILE_PATH
        self.old_stats_configuration = (
            app_module.stats_store.db_path,
            app_module.stats_store.user_root,
        )
        self.user = app_module.User("stats-operator", "")
        app_module.user_manager.users[self.user.id] = self.user
        app_module.Config.SDL_FILE_PATH = str(self.root / "Slide_Digitization_Log.xlsx")
        app_module.stats_store.configure(
            str(self.root / "statistics.sqlite3"), str(self.root / "users")
        )
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = app_module.Config.SDL_SHEET_NAME
        worksheet.append(app_module.SDL_HEADERS)
        workbook.save(app_module.Config.SDL_FILE_PATH)
        workbook.close()
        app_module.app.config.update(TESTING=True, SECRET_KEY="statistics-test")
        self.client = app_module.app.test_client()
        with self.client.session_transaction() as session:
            session["_user_id"] = self.user.id
            session["_fresh"] = True

    def tearDown(self):
        app_module.user_manager.users = self.old_users
        app_module.Config.SDL_FILE_PATH = self.old_sdl_path
        app_module.stats_store.configure(*self.old_stats_configuration)
        self.temporary.cleanup()

    @staticmethod
    def form_data(action="add"):
        return {
            "action": action,
            "accession_id": "NP26-100",
            "organ": "BRAIN",
            "type": "PROSP",
            "slides_count": "2",
            "scanner": "RSCH1 (SS12797)",
            "carousel_rack": "1",
            "date_loaded": "2026-08-25",
            "time_loaded": "09:00",
            "date_unloaded": "",
            "time_unloaded": "",
            "notes": "",
        }

    def test_only_manual_add_increments_accessions(self):
        response = self.client.post("/sdl", data=self.form_data())
        self.assertEqual(response.status_code, 302)
        self.assertEqual(self.user.slides_completed, 0)
        today = datetime.date.today().isoformat()
        self.assertEqual(
            app_module.stats_store._database_rows(self.user.id)[today]["accessions_logged"],
            1,
        )

        workbook = load_workbook(app_module.Config.SDL_FILE_PATH)
        worksheet = workbook[app_module.Config.SDL_SHEET_NAME]
        signature = app_module._sdl_row_signature(worksheet, 2)
        workbook.close()
        update = self.form_data("update")
        update.update({"worksheet_row": "2", "row_signature": signature, "notes": "edit"})
        response = self.client.post("/sdl", data=update)
        self.assertEqual(response.status_code, 302)
        self.assertEqual(
            app_module.stats_store._database_rows(self.user.id)[today]["accessions_logged"],
            1,
        )

    def test_dashboard_heartbeat_and_lifetime_download(self):
        home = self.client.get("/")
        self.assertEqual(home.status_code, 200)
        self.assertIn(b"Your activity", home.data)
        self.assertIn(b"Application navigation", home.data)

        heartbeat = self.client.post("/statistics/heartbeat")
        self.assertEqual(heartbeat.status_code, 200)
        self.assertIn("statistics", heartbeat.get_json())

        yesterday = datetime.date.today() - datetime.timedelta(days=1)
        app_module.stats_store.increment(
            self.user.id, "slides_completed", 2, yesterday
        )
        app_module.stats_store.rollup_user(self.user.id, yesterday)
        page = self.client.get("/statistics/lifetime")
        download = self.client.get("/statistics/lifetime.csv")
        self.assertEqual(page.status_code, 200)
        self.assertIn(yesterday.isoformat().encode(), page.data)
        self.assertEqual(download.status_code, 200)
        self.assertIn("attachment", download.headers["Content-Disposition"])
        download.close()


class AdminStatisticsIntegrationTests(unittest.TestCase):
    def setUp(self):
        self.temporary = tempfile.TemporaryDirectory()
        self.root = Path(self.temporary.name)
        self.old_users = app_module.user_manager.users.copy()
        self.old_stats_configuration = (
            app_module.stats_store.db_path,
            app_module.stats_store.user_root,
        )
        self.admin = app_module.User("stats-admin", "", is_admin=True)
        self.user = app_module.User("team/operator", "")
        app_module.user_manager.users = {
            self.admin.id: self.admin,
            self.user.id: self.user,
        }
        app_module.stats_store.configure(
            str(self.root / "statistics.sqlite3"), str(self.root / "users")
        )
        self.activity_date = datetime.date.today() - datetime.timedelta(days=1)
        app_module.stats_store.increment(
            self.user.id, "slides_completed", 4, self.activity_date
        )
        app_module.stats_store.increment(
            self.user.id, "accessions_logged", 2, self.activity_date
        )
        app_module.stats_store.rollup_user(self.user.id, self.activity_date)
        app_module.app.config.update(TESTING=True, SECRET_KEY="admin-statistics-test")
        self.client = app_module.app.test_client()
        self.login(self.admin)

    def tearDown(self):
        app_module.user_manager.users = self.old_users
        app_module.stats_store.configure(*self.old_stats_configuration)
        self.temporary.cleanup()

    def login(self, user):
        with self.client.session_transaction() as session:
            session.clear()
            session["_user_id"] = user.id
            session["_fresh"] = True

    def test_admin_can_open_user_dashboard_from_user_management(self):
        users_page = self.client.get("/users")
        dashboard = self.client.get(
            "/admin/statistics", query_string={"user_id": self.user.id}
        )

        self.assertEqual(users_page.status_code, 200)
        self.assertIn(b"View Stats", users_page.data)
        self.assertIn(b"user_id=team/operator", users_page.data)
        self.assertEqual(dashboard.status_code, 200)
        self.assertIn(b"team/operator activity", dashboard.data)
        self.assertIn(self.activity_date.isoformat().encode(), dashboard.data)
        self.assertIn(b"View Lifetime Statistics", dashboard.data)
        self.assertNotIn(
            b"window.addEventListener('statistics-update'", dashboard.data
        )

    def test_admin_can_view_and_download_selected_user_lifetime_stats(self):
        query = {"user_id": self.user.id}
        page = self.client.get("/admin/statistics/lifetime", query_string=query)
        download = self.client.get(
            "/admin/statistics/lifetime.csv", query_string=query
        )

        self.assertEqual(page.status_code, 200)
        self.assertIn(b"team/operator", page.data)
        self.assertIn(self.activity_date.isoformat().encode(), page.data)
        self.assertIn(b">4</td>", page.data)
        self.assertEqual(download.status_code, 200)
        self.assertIn("attachment", download.headers["Content-Disposition"])
        self.assertIn(self.activity_date.isoformat().encode(), download.data)
        self.assertIn(b",4,2,", download.data)
        download.close()

    def test_non_admin_cannot_access_selected_user_statistics(self):
        self.login(self.user)
        query = {"user_id": self.user.id}

        for path in (
            "/admin/statistics",
            "/admin/statistics/lifetime",
            "/admin/statistics/lifetime.csv",
        ):
            with self.subTest(path=path):
                response = self.client.get(path, query_string=query)
                self.assertEqual(response.status_code, 403)

    def test_admin_statistics_reject_missing_or_unknown_user(self):
        for path in (
            "/admin/statistics",
            "/admin/statistics/lifetime",
            "/admin/statistics/lifetime.csv",
        ):
            with self.subTest(path=path, user="missing"):
                self.assertEqual(self.client.get(path).status_code, 404)
            with self.subTest(path=path, user="unknown"):
                response = self.client.get(
                    path, query_string={"user_id": "unknown"}
                )
                self.assertEqual(response.status_code, 404)

    def test_admin_statistics_requires_login(self):
        with self.client.session_transaction() as session:
            session.clear()
        response = self.client.get(
            "/admin/statistics", query_string={"user_id": self.user.id}
        )
        self.assertEqual(response.status_code, 302)
        self.assertIn("/login", response.headers["Location"])


class QCStatisticsIntegrationTests(unittest.TestCase):
    def setUp(self):
        self.temporary = tempfile.TemporaryDirectory()
        self.root = Path(self.temporary.name)
        self.batch_base = self.root / "batches"
        self.batch = self.batch_base / "SS100" / "batch-1"
        (self.batch / "label").mkdir(parents=True)
        (self.batch / "macro").mkdir()
        with open(self.batch / "enriched.csv", "w", newline="", encoding="utf-8") as handle:
            writer = csv.DictWriter(
                handle,
                fieldnames=("AccessionID", "BlockNumber", "Stain", "ParsingQCPassed", "original_slide_path"),
            )
            writer.writeheader()
            writer.writerow(
                {
                    "AccessionID": "A12-123",
                    "BlockNumber": "A1",
                    "Stain": "HE",
                    "ParsingQCPassed": "FALSE",
                    "original_slide_path": "slide.svs",
                }
            )
        self.old_values = {
            "LABEL_CHECK_BATCHES": app_module.Config.LABEL_CHECK_BATCHES,
            "INSTANCE_DIR": app_module.Config.INSTANCE_DIR,
            "BACKUP_DIR": app_module.Config.BACKUP_DIR,
        }
        self.old_users = app_module.user_manager.users.copy()
        self.old_stats_configuration = (app_module.stats_store.db_path, app_module.stats_store.user_root)
        app_module.Config.LABEL_CHECK_BATCHES = str(self.batch_base)
        app_module.Config.INSTANCE_DIR = str(self.root / "instance")
        app_module.Config.BACKUP_DIR = str(self.root / "backups")
        app_module.stats_store.configure(str(self.root / "statistics.sqlite3"), str(self.root / "users"))
        app_module.batch_contexts.clear()
        self.user = app_module.User("qc-stats-operator", "")
        app_module.user_manager.users[self.user.id] = self.user
        app_module.app.config.update(TESTING=True, SECRET_KEY="qc-statistics-test")
        self.client = app_module.app.test_client()
        with self.client.session_transaction() as session:
            session["_user_id"] = self.user.id
            session["_fresh"] = True
        batches, _ = app_module.discover_batches()
        self.batch_id = batches[0].id

    def tearDown(self):
        app_module.batch_contexts.clear()
        app_module.user_manager.users = self.old_users
        for key, value in self.old_values.items():
            setattr(app_module.Config, key, value)
        app_module.stats_store.configure(*self.old_stats_configuration)
        self.temporary.cleanup()

    def test_only_marking_complete_increments_slides(self):
        self.client.get(f"/qc?batch={self.batch_id}")
        ordinary_save = {
            "batch": self.batch_id,
            "original_index": "0",
            "accession_id": "A12-123",
            "block_number": "A1",
            "stain": "H-E",
            "action": "save",
        }
        response = self.client.post(f"/update?batch={self.batch_id}", data=ordinary_save)
        self.assertEqual(response.status_code, 302)
        today = datetime.date.today().isoformat()
        self.assertEqual(
            app_module.stats_store._database_rows(self.user.id)[today]["slides_completed"],
            0,
        )

        completion = {**ordinary_save, "action": "next", "complete": "on"}
        with mock.patch.object(app_module, "_start_renaming_job"):
            response = self.client.post(f"/update?batch={self.batch_id}", data=completion)
        self.assertEqual(response.status_code, 302)
        self.assertEqual(
            app_module.stats_store._database_rows(self.user.id)[today]["slides_completed"],
            1,
        )


if __name__ == "__main__":
    unittest.main()
