import tempfile
import unittest
from pathlib import Path
import sys

from openpyxl import Workbook, load_workbook


SRC_DIR = Path(__file__).resolve().parents[1] / "src"
sys.path.insert(0, str(SRC_DIR))

import app as app_module
import renaming


def make_workbook(path: Path, headers):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = app_module.Config.SDL_SHEET_NAME
    worksheet.append(headers)
    workbook.save(path)
    workbook.close()


def mapping_row(accession, organ, original_path, section):
    row = {
        "AccessionID": accession,
        "Organ": organ,
        "PID": "AAAAAA",
        "AccessionDate": "20250101",
        "Timepoint": "XXXX",
        "Stain": "HE",
        "ImageType": "WSI",
        "SampAcqType": "RE",
        "BlockNumber": "A1",
        "SectionCount": section,
        "OriginalPath": original_path,
        "Approved": "True",
    }
    row["NewName"] = renaming.build_new_name(row)
    return row


class SDLPostQCTests(unittest.TestCase):
    def setUp(self):
        self.temporary = tempfile.TemporaryDirectory()
        self.root = Path(self.temporary.name)
        self.batch_base = self.root / "batches"
        self.batch = self.batch_base / "SS12797" / "not-a-date"
        self.batch.mkdir(parents=True)
        self.clone = self.root / "clone"
        self.workbook_path = self.root / "Slide_Digitization_Log.xlsx"
        self.old_sdl = app_module.Config.SDL_FILE_PATH
        self.old_clone = app_module.Config.COPATH_CLONE
        app_module.Config.SDL_FILE_PATH = str(self.workbook_path)
        app_module.Config.COPATH_CLONE = str(self.clone)

    def tearDown(self):
        app_module.Config.SDL_FILE_PATH = self.old_sdl
        app_module.Config.COPATH_CLONE = self.old_clone
        self.temporary.cleanup()

    def test_extra_headers_are_ignored_and_preserved(self):
        headers = [
            "External Before",
            *app_module.SDL_HEADERS[:5],
            "External Middle",
            *app_module.SDL_HEADERS[5:],
            "External After",
        ]
        make_workbook(self.workbook_path, headers)
        workbook = load_workbook(self.workbook_path)
        worksheet = workbook[app_module.Config.SDL_SHEET_NAME]
        columns = {
            worksheet.cell(row=1, column=column).value: column
            for column in range(1, worksheet.max_column + 1)
        }
        worksheet.cell(row=2, column=columns["Accession ID"]).value = "NP25-999"
        worksheet.cell(row=2, column=columns["External Before"]).value = "keep-before"
        worksheet.cell(row=2, column=columns["External Middle"]).value = "keep-middle"
        worksheet.cell(row=2, column=columns["External After"]).value = "keep-after"
        original_signature = app_module._sdl_row_signature(worksheet, 2)
        worksheet.cell(row=2, column=columns["External After"]).value = "changed-extra"
        self.assertEqual(
            original_signature,
            app_module._sdl_row_signature(worksheet, 2),
        )
        worksheet.cell(row=2, column=columns["External After"]).value = "keep-after"
        workbook.save(self.workbook_path)
        workbook.close()
        renaming.atomic_write(
            self.batch / "name_mapping.csv",
            renaming.MAPPING_FIELDS,
            [
                mapping_row(
                    "NP25-100",
                    "TESTIS",
                    r"D:\scanner\2026-07-20\one.svs",
                    "000",
                )
            ],
        )

        added = app_module._update_sdl_after_renaming(self.batch)

        self.assertEqual(1, added)
        workbook = load_workbook(self.workbook_path)
        worksheet = workbook[app_module.Config.SDL_SHEET_NAME]
        columns = app_module._sdl_header_columns(worksheet)
        self.assertEqual(
            "TESTIS", worksheet.cell(row=3, column=columns["Organ"]).value
        )
        self.assertEqual(
            "RSCH1 (SS12797)",
            worksheet.cell(row=3, column=columns["Scanner"]).value,
        )
        self.assertEqual(
            "keep-before",
            worksheet.cell(row=2, column=1).value,
        )
        external_columns = {
            worksheet.cell(row=1, column=column).value: column
            for column in range(1, worksheet.max_column + 1)
        }
        self.assertEqual(
            "keep-middle",
            worksheet.cell(
                row=2, column=external_columns["External Middle"]
            ).value,
        )
        self.assertEqual(
            "keep-after",
            worksheet.cell(
                row=2, column=external_columns["External After"]
            ).value,
        )
        self.assertIsNone(
            worksheet.cell(
                row=3, column=external_columns["External Middle"]
            ).value
        )
        workbook.close()

    def test_existing_accession_is_skipped_case_insensitively(self):
        make_workbook(self.workbook_path, app_module.SDL_HEADERS)
        workbook = load_workbook(self.workbook_path)
        worksheet = workbook[app_module.Config.SDL_SHEET_NAME]
        columns = app_module._sdl_header_columns(worksheet)
        worksheet.cell(row=2, column=columns["Accession ID"]).value = " np25-100 "
        workbook.save(self.workbook_path)
        workbook.close()
        renaming.atomic_write(
            self.batch / "name_mapping.csv",
            renaming.MAPPING_FIELDS,
            [mapping_row("NP25-100", "BRAIN", "one.svs", "000")],
        )

        self.assertEqual(0, app_module._update_sdl_after_renaming(self.batch))

        workbook = load_workbook(self.workbook_path)
        self.assertEqual(
            2, workbook[app_module.Config.SDL_SHEET_NAME].max_row
        )
        workbook.close()

    def test_source_parent_dates_group_counts_and_unknowns(self):
        make_workbook(self.workbook_path, app_module.SDL_HEADERS)
        renaming.atomic_write(
            self.batch / "name_mapping.csv",
            renaming.MAPPING_FIELDS,
            [
                mapping_row(
                    "NP25-100",
                    "",
                    r"D:\scanner\2026-07-20\one.svs",
                    "000",
                ),
                mapping_row(
                    "NP25-100",
                    "",
                    r"D:\scanner\2026-07-20\two.svs",
                    "001",
                ),
                mapping_row(
                    "NP25-100",
                    "",
                    "/scanner/2026-07-21/three.svs",
                    "002",
                ),
                mapping_row(
                    "NP25-100",
                    "",
                    r"D:\scanner\misc\four.svs",
                    "003",
                ),
                mapping_row(
                    "NP25-100",
                    "",
                    "/scanner/also-misc/five.svs",
                    "004",
                ),
            ],
        )
        renaming.atomic_write(
            self.clone / "all_accessions.csv",
            ("AccessionID", "Organ"),
            [{"AccessionID": "NP25-100", "Organ": "BREAST"}],
        )

        self.assertEqual(3, app_module._update_sdl_after_renaming(self.batch))

        workbook = load_workbook(self.workbook_path)
        worksheet = workbook[app_module.Config.SDL_SHEET_NAME]
        rows = app_module._read_sdl_rows(worksheet)
        self.assertEqual(
            [
                ("2026-07-20", "2"),
                ("2026-07-21", "1"),
                (app_module.SDL_UNKNOWN_DATE, "2"),
            ],
            [
                (
                    row["values"]["Date Loaded"],
                    row["values"]["Slides Count"],
                )
                for row in rows
            ],
        )
        self.assertTrue(all(row["values"]["Organ"] == "BREAST" for row in rows))
        self.assertTrue(
            all(row["statuses"]["Ran Label-Check"] for row in rows)
        )
        self.assertTrue(all(row["statuses"]["Finished QC"] for row in rows))
        self.assertTrue(
            all(row["statuses"]["Collected CoPath Data"] for row in rows)
        )
        self.assertTrue(all(row["statuses"]["Renamed"] for row in rows))
        self.assertTrue(
            all(not row["statuses"]["Pushed to SFTP Server"] for row in rows)
        )
        workbook.close()

    def test_batch_date_overrides_source_dates_and_mapping_organ_wins(self):
        self.batch = self.batch.parent / "2026-07-22"
        self.batch.mkdir()
        make_workbook(self.workbook_path, app_module.SDL_HEADERS)
        renaming.atomic_write(
            self.batch / "name_mapping.csv",
            renaming.MAPPING_FIELDS,
            [
                mapping_row("NP25-100", "TESTIS", "/misc/one.svs", "000"),
                mapping_row("NP25-100", "TESTIS", "/misc/two.svs", "001"),
            ],
        )
        renaming.atomic_write(
            self.clone / "all_accessions.csv",
            ("AccessionID", "Organ"),
            [{"AccessionID": "NP25-100", "Organ": "BRAIN"}],
        )

        self.assertEqual(1, app_module._update_sdl_after_renaming(self.batch))

        workbook = load_workbook(self.workbook_path)
        rows = app_module._read_sdl_rows(
            workbook[app_module.Config.SDL_SHEET_NAME]
        )
        self.assertEqual("2026-07-22", rows[0]["values"]["Date Loaded"])
        self.assertEqual("2", rows[0]["values"]["Slides Count"])
        self.assertEqual("TESTIS", rows[0]["values"]["Organ"])
        workbook.close()

    def test_missing_organ_defaults_to_unknown(self):
        make_workbook(self.workbook_path, app_module.SDL_HEADERS)
        renaming.atomic_write(
            self.batch / "name_mapping.csv",
            renaming.MAPPING_FIELDS,
            [mapping_row("NP25-100", "", "/misc/one.svs", "000")],
        )

        app_module._update_sdl_after_renaming(self.batch)

        workbook = load_workbook(self.workbook_path)
        rows = app_module._read_sdl_rows(
            workbook[app_module.Config.SDL_SHEET_NAME]
        )
        self.assertEqual("UNKNOWN", rows[0]["values"]["Organ"])
        workbook.close()

    def test_missing_and_duplicate_required_headers_are_rejected(self):
        make_workbook(self.workbook_path, app_module.SDL_HEADERS[:-1])
        with self.assertRaisesRegex(app_module.SDLWorkbookError, "missing"):
            app_module._load_sdl_workbook()

        make_workbook(
            self.workbook_path,
            (*app_module.SDL_HEADERS, "Accession ID"),
        )
        with self.assertRaisesRegex(app_module.SDLWorkbookError, "duplicated"):
            app_module._load_sdl_workbook()

    def test_date_sentinel_is_valid_for_loaded_and_unloaded(self):
        values = {
            "Accession ID": "NP25-100",
            "Organ": "TESTIS",
            "Type": "SEMINOMA",
            "Slides Count": "1",
            "Scanner": "-----",
            "Carousel Rack": "1",
            "Date Loaded": app_module.SDL_UNKNOWN_DATE,
            "Time Loaded": "12:00",
            "Date Unloaded": app_module.SDL_UNKNOWN_DATE,
            "Time Unloaded": "13:00",
            "Notes": "",
        }

        normalized = app_module._validate_sdl_form(values)

        self.assertEqual(
            app_module.SDL_UNKNOWN_DATE, normalized["Date Loaded"]
        )
        self.assertEqual(
            app_module.SDL_UNKNOWN_DATE, normalized["Date Unloaded"]
        )


if __name__ == "__main__":
    unittest.main()
