import csv
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook


SRC_DIR = Path(__file__).resolve().parents[1] / "src"
sys.path.insert(0, str(SRC_DIR))

import renaming
import app as app_module


def write_csv(path, fields, rows):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)


class RenamingDataTests(unittest.TestCase):
    def setUp(self):
        self.temporary = tempfile.TemporaryDirectory()
        self.root = Path(self.temporary.name)
        self.batch_base = self.root / "batches"
        self.batch = self.batch_base / "SS100" / "batch-1"
        self.clone = self.root / "clone"
        self.batch.mkdir(parents=True)
        write_csv(
            self.batch / "enriched.csv",
            ["AccessionID", "Stain", "BlockNumber", "original_slide_path"],
            [
                {"AccessionID": "NP25-100", "Stain": "HE", "BlockNumber": "B4", "original_slide_path": "one.svs"},
                {"AccessionID": "NP25-100", "Stain": "HE", "BlockNumber": "B4", "original_slide_path": "two.svs"},
            ],
        )
        write_csv(self.clone / "all_accessions.csv", ["AccessionID", "Organ"], [])
        write_csv(
            self.clone / "BRAIN" / "copath_data.csv",
            ["accession_id", "mrn", "PID", "organ"],
            [{"accession_id": "NP24-1", "mrn": "MRN1", "PID": "AAAAAZ", "organ": "BRAIN"}],
        )

    def tearDown(self):
        self.temporary.cleanup()

    def query(self, _batch, accessions, output):
        self.assertEqual(["NP25-100"], list(accessions))
        write_csv(
            output,
            ["accession_id", "mrn", "accession_date", "sample_acquisition_type", "final_diagnosis"],
            [{
                "accession_id": "NP25-100", "mrn": "MRN1",
                "accession_date": "2025-03-04 00:00:00.000",
                "sample_acquisition_type": "Brain resection", "final_diagnosis": "Example",
            }],
        )

    def test_prepare_reuses_mrn_pid_and_builds_unique_names(self):
        renaming.prepare_batch(self.batch, self.clone, self.batch_base, self.query)

        fields, rows = renaming.read_csv(self.batch / "name_mapping.csv")
        self.assertEqual(list(renaming.MAPPING_FIELDS), fields)
        self.assertEqual(["000", "001"], [row["SectionCount"] for row in rows])
        self.assertTrue(all(row["PID"] == "AAAAAZ" for row in rows))
        self.assertEqual("BRAIN_AAAAAZ_20250304_XXXX_HE_WSI_REB4000.svs", rows[0]["NewName"])
        self.assertTrue(all(row["Approved"] == "False" for row in rows))

    def test_prepare_maps_new_accessions_by_mrn_and_organ(self):
        write_csv(
            self.batch / "enriched.csv",
            ["AccessionID", "Stain", "BlockNumber", "original_slide_path"],
            [
                {"AccessionID": "NP25-100", "Stain": "HE", "BlockNumber": "A1", "original_slide_path": "one.svs"},
                {"AccessionID": "NP25-200", "Stain": "HE", "BlockNumber": "A2", "original_slide_path": "two.svs"},
            ],
        )

        def same_patient(_batch, accessions, output):
            self.assertEqual(["NP25-100", "NP25-200"], list(accessions))
            write_csv(
                output,
                ["accession_id", "mrn", "sample_acquisition_type"],
                [
                    {"accession_id": accession, "mrn": "MRN2", "sample_acquisition_type": "Brain resection"}
                    for accession in accessions
                ],
            )

        renaming.prepare_batch(
            self.batch, self.clone, self.batch_base, same_patient
        )

        _, rows = renaming.read_csv(self.batch / "name_mapping.csv")
        self.assertEqual({"NP25-100", "NP25-200"}, {
            row["AccessionID"] for row in rows
        })
        self.assertEqual({"AAAABA"}, {row["PID"] for row in rows})

    def test_prepare_reuses_pair_from_another_staged_batch(self):
        staged_batch = self.batch_base / "SS200" / "batch-2"
        staged = {
            "AccessionID": "NP25-200", "Organ": "BRAIN", "PID": "AAAABA",
            "AccessionDate": "20250101", "Timepoint": "XXXX", "Stain": "HE",
            "ImageType": "WSI", "SampAcqType": "RE", "BlockNumber": "A1",
            "SectionCount": "000", "OriginalPath": "staged.svs", "Approved": "False",
        }
        staged["NewName"] = renaming.build_new_name(staged)
        renaming.atomic_write(
            staged_batch / "name_mapping.csv", renaming.MAPPING_FIELDS, [staged]
        )
        write_csv(
            staged_batch / "pending_CoPath_data.csv",
            ["accession_id", "mrn"],
            [{"accession_id": "NP25-200", "mrn": "MRN2"}],
        )

        def same_patient(_batch, _accessions, output):
            write_csv(
                output,
                ["accession_id", "mrn", "sample_acquisition_type"],
                [{"accession_id": "NP25-100", "mrn": "MRN2", "sample_acquisition_type": "Brain resection"}],
            )

        renaming.prepare_batch(
            self.batch, self.clone, self.batch_base, same_patient
        )

        _, rows = renaming.read_csv(self.batch / "name_mapping.csv")
        self.assertEqual({"AAAABA"}, {row["PID"] for row in rows})

    def test_same_mrn_uses_separate_pid_mapping_for_each_organ(self):
        write_csv(
            self.batch / "enriched.csv",
            ["AccessionID", "Stain", "BlockNumber", "original_slide_path"],
            [
                {"AccessionID": "NP25-100", "Stain": "HE", "BlockNumber": "A1", "original_slide_path": "one.svs"},
                {"AccessionID": "NP25-200", "Stain": "HE", "BlockNumber": "A2", "original_slide_path": "two.svs"},
            ],
        )

        def two_organs(_batch, _accessions, output):
            write_csv(
                output,
                ["accession_id", "mrn", "sample_acquisition_type"],
                [
                    {"accession_id": "NP25-100", "mrn": "MRN2", "sample_acquisition_type": "Brain resection"},
                    {"accession_id": "NP25-200", "mrn": "MRN2", "sample_acquisition_type": "Breast biopsy"},
                ],
            )

        renaming.prepare_batch(
            self.batch, self.clone, self.batch_base, two_organs
        )

        _, rows = renaming.read_csv(self.batch / "name_mapping.csv")
        assignments = {row["Organ"]: row["PID"] for row in rows}
        self.assertEqual({"BRAIN": "AAAABA", "BREAST": "AAAAAA"}, assignments)

    def test_prepare_missing_report_uses_unknown_and_index_only_on_finalize(self):
        def no_results(_batch, _accessions, output):
            write_csv(output, ["accession_id"], [])

        renaming.prepare_batch(self.batch, self.clone, self.batch_base, no_results)
        _, rows = renaming.read_csv(self.batch / "name_mapping.csv")
        self.assertEqual("UNKNOWN", rows[0]["Organ"])
        self.assertEqual("AAAAAA", rows[0]["PID"])
        for row in rows:
            row["Approved"] = "True"
        renaming.atomic_write(self.batch / "name_mapping.csv", renaming.MAPPING_FIELDS, rows)

        renaming.finalize_batch(self.batch, self.clone)

        _, index_rows = renaming.read_csv(self.clone / "all_iuh_identifiers.csv")
        self.assertIn({
            "AccessionID": "NP25-100", "Organ": "UNKNOWN", "MRN": "", "PID": "AAAAAA",
        }, index_rows)
        unknown_path = self.clone / "UNKNOWN" / "copath_data.csv"
        self.assertTrue(unknown_path.exists())
        headers, unknown_rows = renaming.read_csv(unknown_path)
        self.assertEqual(list(renaming.COPATH_FIELDS), headers)
        self.assertEqual([], unknown_rows)

    def test_organ_change_pid_follows_highest_staged_pid(self):
        renaming.prepare_batch(self.batch, self.clone, self.batch_base, self.query)
        _, mapping = renaming.read_csv(self.batch / "name_mapping.csv")
        for row in mapping:
            row.update({"Organ": "OTHER", "PID": "AAAAAA"})
            row["NewName"] = renaming.build_new_name(row)
        renaming.atomic_write(
            self.batch / "name_mapping.csv", renaming.MAPPING_FIELDS, mapping
        )
        write_csv(
            self.batch / "pending_CoPath_data.csv",
            ["accession_id", "mrn"],
            [{"accession_id": "NP25-100", "mrn": "MRN2"}],
        )
        staged_batch = self.batch_base / "SS200" / "batch-2"
        staged = dict(mapping[0])
        staged.update({
            "AccessionID": "NP25-200", "Organ": "BRAIN", "PID": "AAAABA"
        })
        staged["NewName"] = renaming.build_new_name(staged)
        renaming.atomic_write(
            staged_batch / "name_mapping.csv", renaming.MAPPING_FIELDS, [staged]
        )

        pid = renaming.pid_after_organ_change(
            self.batch, self.clone, self.batch_base, "NP25-100", "BRAIN"
        )
        reserved_pid = renaming.pid_after_organ_change(
            self.batch, self.clone, self.batch_base, "NP25-100", "BRAIN",
            ["AAAABB"],
        )

        self.assertEqual("AAAABB", pid)
        self.assertEqual("AAAABC", reserved_pid)

    def test_organ_change_pid_reuses_existing_mrn_organ_pid(self):
        renaming.prepare_batch(self.batch, self.clone, self.batch_base, self.query)
        _, mapping = renaming.read_csv(self.batch / "name_mapping.csv")
        for row in mapping:
            row.update({"Organ": "OTHER", "PID": "AAAAAA"})
            row["NewName"] = renaming.build_new_name(row)
        renaming.atomic_write(
            self.batch / "name_mapping.csv", renaming.MAPPING_FIELDS, mapping
        )

        pid = renaming.pid_after_organ_change(
            self.batch, self.clone, self.batch_base, "NP25-100", "BRAIN"
        )

        self.assertEqual("AAAAAZ", pid)

    def test_empty_clone_is_initialized_with_all_required_csvs(self):
        empty_clone = self.root / "empty-clone"

        accession_org, rows_by_organ, headers_by_organ = renaming._clone_rows(empty_clone)

        self.assertEqual({}, accession_org)
        index_headers, index_rows = renaming.read_csv(empty_clone / "all_iuh_identifiers.csv")
        self.assertEqual(list(renaming.IDENTIFIER_FIELDS), index_headers)
        self.assertEqual([], index_rows)
        for organ in renaming.ORGANS:
            self.assertEqual([], rows_by_organ[organ])
            self.assertEqual(list(renaming.COPATH_FIELDS), headers_by_organ[organ])
            self.assertTrue((empty_clone / organ / "copath_data.csv").is_file())

    def test_legacy_identifier_index_is_merged_and_recoverably_retired(self):
        renaming.initialize_clone(self.clone)

        self.assertFalse((self.clone / "all_accessions.csv").exists())
        self.assertTrue(
            (self.clone / "migration_backups" / "identifier_indexes" / "all_accessions.csv").exists()
        )
        fields, rows = renaming.read_csv(self.clone / "all_iuh_identifiers.csv")
        self.assertEqual(list(renaming.IDENTIFIER_FIELDS), fields)
        self.assertIn({
            "AccessionID": "NP24-1", "Organ": "BRAIN", "MRN": "MRN1", "PID": "AAAAAZ",
        }, rows)

    def test_longitudinal_history_is_staged_logged_and_committed_after_approval(self):
        def initial(_batch, accessions, output):
            self.assertEqual(["NP25-100"], list(accessions))
            write_csv(output, ["accession_id", "mrn", "accession_date", "sample_acquisition_type", "report"], [{
                "accession_id": "NP25-100", "mrn": "MRN2", "accession_date": "2025-03-04",
                "sample_acquisition_type": "Brain resection", "report": "Seed",
            }])

        def history(_batch, accessions, output):
            self.assertEqual(["NP25-100"], list(accessions))
            write_csv(output, ["accession_id", "mrn", "accession_date", "sample_acquisition_type", "report"], [
                {"accession_id": "NP25-100", "mrn": "MRN2", "accession_date": "2025-03-04", "sample_acquisition_type": "Brain resection", "report": "Seed"},
                {"accession_id": "SP20-5", "mrn": "MRN2", "accession_date": "2020-01-02", "sample_acquisition_type": "Breast biopsy", "report": "History"},
            ])

        renaming.prepare_batch(self.batch, self.clone, self.batch_base, initial)
        self.assertEqual("pending", renaming.read_history_job(self.batch)["status"])
        renaming.stage_longitudinal_history(
            self.batch, self.clone, self.batch_base, history
        )
        _, staged = renaming.read_csv(self.batch / "pending_CoPath_history.csv")
        breast = next(row for row in staged if row["accession_id"] == "SP20-5")
        self.assertEqual(("BREAST", "BP", "20200102"), (
            breast["organ"], breast["_sampacqtype"], breast["_accdate"],
        ))
        _, log = renaming.read_csv(self.batch / "copath_longitudinal_jobs.csv")
        self.assertEqual("NONE", log[0]["Error"])

        _, mapping = renaming.read_csv(self.batch / "name_mapping.csv")
        for row in mapping:
            row["Approved"] = "True"
        renaming.atomic_write(self.batch / "name_mapping.csv", renaming.MAPPING_FIELDS, mapping)
        renaming.finalize_batch(self.batch, self.clone)

        _, identifiers = renaming.read_csv(self.clone / "all_iuh_identifiers.csv")
        self.assertEqual({"NP25-100", "SP20-5", "NP24-1"}, {
            row["AccessionID"] for row in identifiers
        })
        self.assertEqual("committed", renaming.read_history_job(self.batch)["status"])

    def test_longitudinal_log_records_error_then_retry(self):
        def initial(_batch, _accessions, output):
            write_csv(output, ["accession_id", "mrn"], [{
                "accession_id": "NP25-100", "mrn": "MRN2",
            }])

        renaming.prepare_batch(self.batch, self.clone, self.batch_base, initial)
        with self.assertRaisesRegex(RuntimeError, "database unavailable"):
            renaming.stage_longitudinal_history(
                self.batch, self.clone, self.batch_base,
                lambda *_args: (_ for _ in ()).throw(RuntimeError("database unavailable")),
            )
        _, failed = renaming.read_csv(self.batch / "copath_longitudinal_jobs.csv")
        self.assertEqual("ERROR: database unavailable", failed[0]["Error"])

        def retry(_batch, _accessions, output):
            write_csv(output, ["accession_id", "mrn"], [{
                "accession_id": "NP25-100", "mrn": "MRN2",
            }])

        renaming.stage_longitudinal_history(self.batch, self.clone, self.batch_base, retry)
        _, retried = renaming.read_csv(self.batch / "copath_longitudinal_jobs.csv")
        self.assertEqual("RETRY", retried[0]["Error"])

    def test_update_group_rejects_stale_signature_and_merges_to_target(self):
        rows = []
        for accession, path, pid in (("NP25-100", "one.svs", "AAAAAA"), ("NP25-200", "two.svs", "AAAAAB")):
            row = {
                "AccessionID": accession, "Organ": "BRAIN", "PID": pid,
                "AccessionDate": "20250101", "Timepoint": "XXXX", "Stain": "HE",
                "ImageType": "WSI", "SampAcqType": "RE", "BlockNumber": "A1",
                "SectionCount": "001", "OriginalPath": path, "Approved": "True",
            }
            row["NewName"] = renaming.build_new_name(row)
            rows.append(row)
        renaming.atomic_write(self.batch / "name_mapping.csv", renaming.MAPPING_FIELDS, rows)
        with self.assertRaisesRegex(renaming.RenamingError, "changed in another session"):
            renaming.update_group(self.batch / "name_mapping.csv", "NP25-100", {}, {}, "bad")

        values = {field: rows[0][field] for field in ("Organ", "PID", "AccessionDate", "Timepoint", "ImageType", "SampAcqType")}
        values["AccessionID"] = "NP25-200"
        updated, merged = renaming.update_group(
            self.batch / "name_mapping.csv", "NP25-100", values,
            {"one.svs": {"Stain": "HE", "BlockNumber": "A2", "SectionCount": "002"}},
            renaming.mapping_signature(rows),
        )
        self.assertTrue(merged)
        self.assertTrue(all(row["AccessionID"] == "NP25-200" for row in updated))
        self.assertTrue(all(row["PID"] == "AAAAAB" for row in updated))
        self.assertTrue(all(row["Approved"] == "False" for row in updated))

    def test_retry_reuses_existing_accession_and_updates_batch_tables(self):
        renaming.prepare_batch(self.batch, self.clone, self.batch_base, self.query)
        pending_path = self.batch / "pending_CoPath_data.csv"
        headers, pending = renaming.read_csv(pending_path)
        corrected = dict(pending[0])
        corrected.update({
            "accession_id": "NP25-200",
            "accession_date": "2025-04-05",
        })
        renaming.atomic_write(pending_path, headers, [*pending, corrected])

        def unexpected_query(_batch, _accessions, _output):
            self.fail("existing corrected accession should not requery CoPath")

        renaming.retry_group(
            self.batch,
            self.clone,
            self.batch_base,
            "NP25-100",
            "NP25-200",
            unexpected_query,
        )

        _, mapping = renaming.read_csv(self.batch / "name_mapping.csv")
        _, enriched = renaming.read_csv(self.batch / "enriched.csv")
        _, pending = renaming.read_csv(pending_path)
        self.assertEqual({"NP25-200"}, {row["AccessionID"] for row in mapping})
        self.assertEqual({"NP25-200"}, {row["AccessionID"] for row in enriched})
        self.assertEqual({"NP25-200"}, {renaming.row_accession(row) for row in pending})

    def test_retry_reuses_accession_from_pending_history(self):
        renaming.prepare_batch(self.batch, self.clone, self.batch_base, self.query)
        history_path = self.batch / "pending_CoPath_history.csv"
        write_csv(
            history_path,
            ["accession_id", "mrn", "accession_date", "sample_acquisition_type"],
            [{
                "accession_id": "NP25-200",
                "mrn": "MRN2",
                "accession_date": "2025-04-05",
                "sample_acquisition_type": "Brain resection",
            }],
        )

        def unexpected_query(_batch, _accessions, _output):
            self.fail("pending history accession should not requery CoPath")

        renaming.retry_group(
            self.batch,
            self.clone,
            self.batch_base,
            "NP25-100",
            "NP25-200",
            unexpected_query,
        )

        _, mapping = renaming.read_csv(self.batch / "name_mapping.csv")
        _, pending = renaming.read_csv(self.batch / "pending_CoPath_data.csv")
        corrected = next(
            row for row in pending if renaming.row_accession(row) == "NP25-200"
        )
        self.assertEqual({"NP25-200"}, {row["AccessionID"] for row in mapping})
        self.assertEqual("MRN2", corrected["mrn"])
        self.assertFalse(history_path.exists())

    def test_retry_merges_typo_group_into_existing_accession(self):
        rows = []
        for accession, path in (
            ("NP25-100", "typo.svs"),
            ("NP25-200", "correct.svs"),
        ):
            row = {
                "AccessionID": accession,
                "Organ": "BRAIN",
                "PID": "AAAAAA",
                "AccessionDate": "20250304",
                "Timepoint": "XXXX",
                "Stain": "HE",
                "ImageType": "WSI",
                "SampAcqType": "RE",
                "BlockNumber": "B4",
                "SectionCount": "000",
                "OriginalPath": path,
                "Approved": "False",
            }
            row["NewName"] = renaming.build_new_name(row)
            rows.append(row)
        renaming.atomic_write(
            self.batch / "name_mapping.csv", renaming.MAPPING_FIELDS, rows
        )
        write_csv(
            self.batch / "enriched.csv",
            ["AccessionID", "Stain", "BlockNumber", "original_slide_path"],
            [
                {"AccessionID": "NP25-100", "Stain": "HE", "BlockNumber": "B4", "original_slide_path": "typo.svs"},
                {"AccessionID": "NP25-200", "Stain": "HE", "BlockNumber": "B4", "original_slide_path": "correct.svs"},
            ],
        )

        def unexpected_query(_batch, _accessions, _output):
            self.fail("existing target accession should not requery CoPath")

        renaming.retry_group(
            self.batch,
            self.clone,
            self.batch_base,
            "NP25-100",
            "NP25-200",
            unexpected_query,
        )

        _, merged = renaming.read_csv(self.batch / "name_mapping.csv")
        groups = renaming.group_mapping(merged, {})
        _, enriched = renaming.read_csv(self.batch / "enriched.csv")
        self.assertEqual(1, len(groups))
        self.assertEqual("NP25-200", groups[0]["accession"])
        self.assertEqual(2, len(groups[0]["slides"]))
        self.assertEqual(["000", "001"], [row["SectionCount"] for row in merged])
        self.assertEqual(2, len({row["NewName"] for row in merged}))
        self.assertEqual({"NP25-200"}, {row["AccessionID"] for row in enriched})


class RenamingPageTests(unittest.TestCase):
    def setUp(self):
        self.temporary = tempfile.TemporaryDirectory()
        self.root = Path(self.temporary.name)
        self.batch_base = self.root / "batches"
        self.batch = self.batch_base / "SS100" / "batch-1"
        self.clone = self.root / "clone"
        (self.batch / "label").mkdir(parents=True)
        (self.batch / "macro").mkdir()
        write_csv(
            self.batch / "enriched.csv",
            ["AccessionID", "Stain", "BlockNumber", "ParsingQCPassed", "original_slide_path"],
            [{"AccessionID": "NP25-100", "Stain": "HE", "BlockNumber": "B4", "ParsingQCPassed": "TRUE", "original_slide_path": "one.svs"}],
        )
        write_csv(self.batch / "completed_stages.csv", ["QC", "Renamed"], [{"QC": "True", "Renamed": "False"}])
        write_csv(self.clone / "all_accessions.csv", ["AccessionID", "Organ"], [])
        row = {
            "AccessionID": "NP25-100", "Organ": "BRAIN", "PID": "AAAAAA",
            "AccessionDate": "20250304", "Timepoint": "XXXX", "Stain": "HE",
            "ImageType": "WSI", "SampAcqType": "RE", "BlockNumber": "B4",
            "SectionCount": "001", "OriginalPath": "one.svs", "Approved": "False",
        }
        row["NewName"] = renaming.build_new_name(row)
        renaming.atomic_write(self.batch / "name_mapping.csv", renaming.MAPPING_FIELDS, [row])
        write_csv(
            self.batch / "pending_CoPath_data.csv",
            ["accession_id", "mrn", "final_diagnosis"],
            [{"accession_id": "NP25-100", "mrn": "MRN1", "final_diagnosis": "Diagnosis text"}],
        )
        self.old_batch_base = app_module.Config.LABEL_CHECK_BATCHES
        self.old_clone = app_module.Config.COPATH_CLONE
        self.old_instance = app_module.Config.INSTANCE_DIR
        self.old_sdl = app_module.Config.SDL_FILE_PATH
        self.old_users = app_module.user_manager.users.copy()
        app_module.Config.LABEL_CHECK_BATCHES = str(self.batch_base)
        app_module.Config.COPATH_CLONE = str(self.clone)
        app_module.Config.INSTANCE_DIR = str(self.root / "instance")
        app_module.Config.SDL_FILE_PATH = str(self.root / "Slide_Digitization_Log.xlsx")
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = app_module.Config.SDL_SHEET_NAME
        worksheet.append(app_module.SDL_HEADERS)
        workbook.save(app_module.Config.SDL_FILE_PATH)
        workbook.close()
        app_module.batch_contexts.clear()
        app_module._renaming_jobs.clear()
        catalog_batches, _ = app_module.discover_batches()
        app_module.batch_catalog.update_stages(
            app_module.Config.INSTANCE_DIR,
            catalog_batches[0].id,
            qc_complete=True,
        )
        self.user = app_module.User("renamer", "", False)
        app_module.user_manager.users[self.user.id] = self.user
        app_module.app.config.update(TESTING=True, SECRET_KEY="renaming-test")
        self.client = app_module.app.test_client()
        with self.client.session_transaction() as session:
            session["_user_id"] = self.user.id
            session["_fresh"] = True

    def add_second_accession(self):
        _, rows = renaming.read_csv(self.batch / "name_mapping.csv")
        second = dict(rows[0])
        second.update({
            "AccessionID": "NP25-200", "PID": "AAAAAB", "OriginalPath": "two.svs",
            "SectionCount": "001",
        })
        second["NewName"] = renaming.build_new_name(second)
        renaming.atomic_write(
            self.batch / "name_mapping.csv", renaming.MAPPING_FIELDS, [rows[0], second]
        )
        return [rows[0], second]

    @staticmethod
    def approval_data(rows, **overrides):
        data = {
            "old_accession": "NP25-100",
            "mapping_signature": renaming.mapping_signature(rows),
            "accession_id": "NP25-100",
            "organ": "BRAIN",
            "pid": "AAAAAA",
            "accession_date": "20250304",
            "timepoint": "XXXX",
            "image_type": "WSI",
            "samp_acq_type": "RE",
            "slide_count": "1",
            "original_path_0": "one.svs",
            "stain_0": "HE",
            "block_number_0": "B4",
            "section_count_0": "001",
        }
        data.update(overrides)
        return data

    def tearDown(self):
        app_module.Config.LABEL_CHECK_BATCHES = self.old_batch_base
        app_module.Config.COPATH_CLONE = self.old_clone
        app_module.Config.INSTANCE_DIR = self.old_instance
        app_module.Config.SDL_FILE_PATH = self.old_sdl
        app_module.user_manager.users = self.old_users
        app_module.batch_contexts.clear()
        app_module._renaming_jobs.clear()
        self.temporary.cleanup()

    def test_page_lists_batch_and_renders_mapping_and_report(self):
        listing = self.client.get("/renaming")
        batches, _ = app_module._renaming_batches()
        detail = self.client.get(f"/renaming?batch={batches[0].id}")

        self.assertEqual(200, listing.status_code)
        self.assertIn(b"SS100/batch-1", listing.data)
        self.assertEqual(200, detail.status_code)
        self.assertIn(b"NP25-100", detail.data)
        self.assertIn(b"Diagnosis text", detail.data)
        self.assertIn(b"fetch(form.action", detail.data)
        self.assertIn(b"pidLookupUrl", detail.data)
        self.assertIn(b"pidLookupQueue", detail.data)
        self.assertIn(b"reserved_pid", detail.data)
        self.assertIn(b'name="pid" value="AAAAAA" maxlength="6" required readonly', detail.data)
        self.assertIn(b'class="success approve-button"', detail.data)

    def test_replace_sdl_accession_updates_existing_rows(self):
        workbook = load_workbook(app_module.Config.SDL_FILE_PATH)
        worksheet = workbook[app_module.Config.SDL_SHEET_NAME]
        old_row = {header: None for header in app_module.SDL_HEADERS}
        old_row["Accession ID"] = "NP25-100"
        other_row = dict(old_row)
        other_row["Accession ID"] = "NP25-300"
        worksheet.append([old_row[header] for header in app_module.SDL_HEADERS])
        worksheet.append([other_row[header] for header in app_module.SDL_HEADERS])
        workbook.save(app_module.Config.SDL_FILE_PATH)
        workbook.close()

        changed = app_module._replace_sdl_accession("NP25-100", "NP25-200")

        self.assertEqual(1, changed)
        workbook = load_workbook(app_module.Config.SDL_FILE_PATH)
        worksheet = workbook[app_module.Config.SDL_SHEET_NAME]
        headers = app_module._sdl_header_columns(worksheet)
        values = [
            worksheet.cell(row=row, column=headers["Accession ID"]).value
            for row in range(2, worksheet.max_row + 1)
        ]
        self.assertEqual(["NP25-200", "NP25-300"], values)
        workbook.close()

    def test_pid_preview_returns_next_staged_pid_for_changed_organ(self):
        _, rows = renaming.read_csv(self.batch / "name_mapping.csv")
        rows[0].update({"Organ": "OTHER", "PID": "AAAAAA"})
        rows[0]["NewName"] = renaming.build_new_name(rows[0])
        renaming.atomic_write(
            self.batch / "name_mapping.csv", renaming.MAPPING_FIELDS, rows
        )
        staged_batch = self.batch_base / "SS200" / "batch-2"
        staged = dict(rows[0])
        staged.update({
            "AccessionID": "NP25-200", "Organ": "BRAIN", "PID": "AAAAAB"
        })
        staged["NewName"] = renaming.build_new_name(staged)
        renaming.atomic_write(
            staged_batch / "name_mapping.csv", renaming.MAPPING_FIELDS, [staged]
        )
        batches, _ = app_module._renaming_batches()

        response = self.client.get(
            f"/renaming/pid/{batches[0].id}",
            query_string={
                "accession": "NP25-100",
                "organ": "BRAIN",
                "mapping_signature": renaming.mapping_signature(rows),
            },
        )

        self.assertEqual(200, response.status_code)
        self.assertEqual({"success": True, "pid": "AAAAAC"}, response.get_json())

        reserved_response = self.client.get(
            f"/renaming/pid/{batches[0].id}",
            query_string=[
                ("accession", "NP25-100"),
                ("organ", "BRAIN"),
                ("mapping_signature", renaming.mapping_signature(rows)),
                ("reserved_pid", "AAAAAC"),
            ],
        )
        self.assertEqual(200, reserved_response.status_code)
        self.assertEqual(
            {"success": True, "pid": "AAAAAD"}, reserved_response.get_json()
        )

    def test_pid_preview_rejects_stale_mapping_signature(self):
        batches, _ = app_module._renaming_batches()

        response = self.client.get(
            f"/renaming/pid/{batches[0].id}",
            query_string={
                "accession": "NP25-100",
                "organ": "OTHER",
                "mapping_signature": "stale",
            },
        )

        self.assertEqual(409, response.status_code)
        self.assertFalse(response.get_json()["success"])
        self.assertIn("changed in another session", response.get_json()["message"])

    def test_approval_overrides_submitted_pid_after_organ_change(self):
        rows = self.add_second_accession()
        rows[0].update({"Organ": "OTHER", "PID": "AAAAAA"})
        rows[0]["NewName"] = renaming.build_new_name(rows[0])
        rows[1].update({"Organ": "BRAIN", "PID": "AAAAAB"})
        rows[1]["NewName"] = renaming.build_new_name(rows[1])
        renaming.atomic_write(
            self.batch / "name_mapping.csv", renaming.MAPPING_FIELDS, rows
        )
        batches, _ = app_module._renaming_batches()

        response = self.client.post(
            f"/renaming/approve/{batches[0].id}",
            data=self.approval_data(rows, organ="BRAIN", pid="ZZZZZZ"),
            headers={"Accept": "application/json"},
        )

        self.assertEqual(200, response.status_code)
        self.assertTrue(response.get_json()["success"])
        _, saved = renaming.read_csv(self.batch / "name_mapping.csv")
        first = next(row for row in saved if row["AccessionID"] == "NP25-100")
        self.assertEqual("BRAIN", first["Organ"])
        self.assertEqual("AAAAAC", first["PID"])

    def test_approval_ignores_submitted_pid_without_organ_change(self):
        _, rows = renaming.read_csv(self.batch / "name_mapping.csv")
        batches, _ = app_module._renaming_batches()

        response = self.client.post(
            f"/renaming/approve/{batches[0].id}",
            data=self.approval_data(rows, pid="ZZZZZZ"),
            headers={"Accept": "application/json"},
        )

        self.assertEqual(200, response.status_code)
        self.assertTrue(response.get_json()["success"])
        _, saved = renaming.read_csv(self.batch / "name_mapping.csv")
        self.assertEqual("AAAAAA", saved[0]["PID"])

    def test_multiple_accessions_render_as_complete_rows_before_expansion(self):
        self.add_second_accession()
        batches, _ = app_module._renaming_batches()

        response = self.client.get(f"/renaming?batch={batches[0].id}")

        self.assertEqual(200, response.status_code)
        self.assertEqual(2, response.data.count(b'class="accession-form'))
        self.assertEqual(2, response.data.count(b'name="accession_id"'))
        self.assertEqual(2, response.data.count(b'class="slide-toggle secondary"'))
        self.assertEqual(2, response.data.count(b'class="expanded"'))
        self.assertNotIn(b'<details>\n                        <summary class="cell"', response.data)

    def test_report_view_is_a_row_button_with_its_own_full_width_panel(self):
        batches, _ = app_module._renaming_batches()

        response = self.client.get(f"/renaming?batch={batches[0].id}")

        self.assertEqual(200, response.status_code)
        self.assertIn(b'class="report-toggle secondary"', response.data)
        self.assertIn(b'>View</button>', response.data)
        self.assertIn(b'class="expanded report-expanded"', response.data)
        self.assertIn(b'id="reports-0" hidden', response.data)
        self.assertNotIn(b'<details class="report">', response.data)

    def test_async_approval_updates_one_group_without_redirecting(self):
        rows = self.add_second_accession()
        batches, _ = app_module._renaming_batches()

        response = self.client.post(
            f"/renaming/approve/{batches[0].id}",
            data=self.approval_data(rows, stain_0="H&E"),
            headers={"Accept": "application/json"},
        )

        self.assertEqual(200, response.status_code)
        payload = response.get_json()
        self.assertTrue(payload["success"])
        self.assertFalse(payload["finalized"])
        self.assertFalse(payload["merged"])
        self.assertIn('class="accession-form approved"', payload["group_html"])
        _, saved = renaming.read_csv(self.batch / "name_mapping.csv")
        first = next(row for row in saved if row["AccessionID"] == "NP25-100")
        second = next(row for row in saved if row["AccessionID"] == "NP25-200")
        self.assertEqual("H&E", first["Stain"])
        self.assertEqual("True", first["Approved"])
        self.assertEqual("False", second["Approved"])
        self.assertEqual(renaming.mapping_signature(saved), payload["signature"])

    def test_async_approval_rejects_a_stale_mapping_without_saving(self):
        rows = self.add_second_accession()
        batches, _ = app_module._renaming_batches()

        response = self.client.post(
            f"/renaming/approve/{batches[0].id}",
            data=self.approval_data(rows, mapping_signature="stale"),
            headers={"Accept": "application/json"},
        )

        self.assertEqual(409, response.status_code)
        payload = response.get_json()
        self.assertFalse(payload["success"])
        self.assertFalse(payload["saved"])
        self.assertIn("changed in another session", payload["message"])
        _, saved = renaming.read_csv(self.batch / "name_mapping.csv")
        self.assertEqual(rows, saved)

    def test_async_merge_returns_one_combined_pending_group(self):
        rows = self.add_second_accession()
        batches, _ = app_module._renaming_batches()

        response = self.client.post(
            f"/renaming/approve/{batches[0].id}",
            data=self.approval_data(rows, accession_id="NP25-200"),
            headers={"Accept": "application/json"},
        )

        self.assertEqual(200, response.status_code)
        payload = response.get_json()
        self.assertTrue(payload["success"])
        self.assertTrue(payload["merged"])
        self.assertEqual("NP25-200", payload["accession"])
        self.assertIn('name="slide_count" value="2"', payload["group_html"])
        self.assertIn("<span>Pending</span>", payload["group_html"])
        _, saved = renaming.read_csv(self.batch / "name_mapping.csv")
        self.assertEqual({"NP25-200"}, {row["AccessionID"] for row in saved})
        self.assertTrue(all(row["Approved"] == "False" for row in saved))

    def test_approval_finalizes_clone_and_completed_stage(self):
        batches, _ = app_module._renaming_batches()
        _, rows = renaming.read_csv(self.batch / "name_mapping.csv")
        response = self.client.post(
            f"/renaming/approve/{batches[0].id}",
            data=self.approval_data(rows),
        )

        self.assertEqual(302, response.status_code)
        catalog_row = app_module.batch_catalog.get_batch(
            app_module.Config.INSTANCE_DIR, batches[0].id
        )
        self.assertTrue(catalog_row["qc_complete"])
        self.assertTrue(catalog_row["renamed_complete"])
        _, index_rows = renaming.read_csv(self.clone / "all_iuh_identifiers.csv")
        self.assertEqual("NP25-100", index_rows[0]["AccessionID"])
        _, clone_rows = renaming.read_csv(self.clone / "BRAIN" / "copath_data.csv")
        self.assertEqual("AAAAAA", clone_rows[0]["PID"])
        workbook = load_workbook(app_module.Config.SDL_FILE_PATH)
        worksheet = workbook[app_module.Config.SDL_SHEET_NAME]
        headers = app_module._sdl_header_columns(worksheet)
        self.assertEqual(
            "NP25-100",
            worksheet.cell(row=2, column=headers["Accession ID"]).value,
        )
        self.assertEqual(
            app_module.SDL_UNKNOWN_DATE,
            worksheet.cell(row=2, column=headers["Date Loaded"]).value,
        )
        self.assertEqual(
            "-----",
            worksheet.cell(row=2, column=headers["Scanner"]).value,
        )
        workbook.close()

    def test_async_final_approval_returns_destination(self):
        batches, _ = app_module._renaming_batches()
        _, rows = renaming.read_csv(self.batch / "name_mapping.csv")

        response = self.client.post(
            f"/renaming/approve/{batches[0].id}",
            data=self.approval_data(rows),
            headers={"Accept": "application/json"},
        )

        self.assertEqual(200, response.status_code)
        payload = response.get_json()
        self.assertTrue(payload["success"])
        self.assertTrue(payload["finalized"])
        self.assertEqual("/renaming", payload["redirect_url"])
        catalog_row = app_module.batch_catalog.get_batch(
            app_module.Config.INSTANCE_DIR, batches[0].id
        )
        self.assertTrue(catalog_row["qc_complete"])
        self.assertTrue(catalog_row["renamed_complete"])

    def test_sdl_failure_leaves_batch_available_for_retry(self):
        Path(app_module.Config.SDL_FILE_PATH).unlink()
        batches, _ = app_module._renaming_batches()
        _, rows = renaming.read_csv(self.batch / "name_mapping.csv")

        response = self.client.post(
            f"/renaming/approve/{batches[0].id}",
            data={
                "old_accession": "NP25-100",
                "mapping_signature": renaming.mapping_signature(rows),
                "accession_id": "NP25-100",
                "organ": "BRAIN",
                "pid": "AAAAAA",
                "accession_date": "20250304",
                "timepoint": "XXXX",
                "image_type": "WSI",
                "samp_acq_type": "RE",
                "slide_count": "1",
                "original_path_0": "one.svs",
                "stain_0": "HE",
                "block_number_0": "B4",
                "section_count_0": "001",
            },
        )

        self.assertEqual(302, response.status_code)
        catalog_row = app_module.batch_catalog.get_batch(
            app_module.Config.INSTANCE_DIR, batches[0].id
        )
        self.assertTrue(catalog_row["qc_complete"])
        self.assertFalse(catalog_row["renamed_complete"])
        self.assertEqual(1, len(app_module._renaming_batches()[0]))

    def test_async_finalization_error_returns_the_saved_group_state(self):
        Path(app_module.Config.SDL_FILE_PATH).unlink()
        batches, _ = app_module._renaming_batches()
        _, rows = renaming.read_csv(self.batch / "name_mapping.csv")

        response = self.client.post(
            f"/renaming/approve/{batches[0].id}",
            data=self.approval_data(rows),
            headers={"Accept": "application/json"},
        )

        self.assertEqual(500, response.status_code)
        payload = response.get_json()
        self.assertFalse(payload["success"])
        self.assertTrue(payload["saved"])
        self.assertIn('class="accession-form approved"', payload["group_html"])
        _, saved = renaming.read_csv(self.batch / "name_mapping.csv")
        self.assertEqual(renaming.mapping_signature(saved), payload["signature"])


if __name__ == "__main__":
    unittest.main()
