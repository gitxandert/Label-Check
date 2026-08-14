import csv
import io
import sys
import tempfile
import unittest
from pathlib import Path
from types import SimpleNamespace
from unittest import mock

from openpyxl import Workbook, load_workbook

SRC_DIR = Path(__file__).resolve().parents[1] / "src"
sys.path.insert(0, str(SRC_DIR))

import app as app_module
import renaming


class FakeProcess:
    def __init__(self, output=b"", return_code=0):
        self.stdout = io.BytesIO(output)
        self.return_code = return_code

    def wait(self):
        return self.return_code

    def poll(self):
        return self.return_code


def write_csv(path, fields, rows):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)


def mapping_row(original_path, pid, section):
    row = {
        "AccessionID": "NP25-100",
        "Organ": "BRAIN",
        "PID": pid,
        "AccessionDate": "20250101",
        "Timepoint": "XXXX",
        "Stain": "HE",
        "ImageType": "WSI",
        "SampAcqType": "RE",
        "BlockNumber": "B4",
        "SectionCount": section,
        "OriginalPath": original_path,
        "Approved": "True",
    }
    row["NewName"] = renaming.build_new_name(row)
    return row


class TQTransferTests(unittest.TestCase):
    def setUp(self):
        self.temporary = tempfile.TemporaryDirectory()
        self.root = Path(self.temporary.name)
        self.batch_base = self.root / "batches"
        self.batch = self.batch_base / "SS12797" / "2026-07-20"
        (self.batch / "label").mkdir(parents=True)
        (self.batch / "macro").mkdir()
        write_csv(
            self.batch / "enriched.csv",
            [
                "AccessionID",
                "Stain",
                "BlockNumber",
                "ParsingQCPassed",
                "original_slide_path",
            ],
            [
                {
                    "AccessionID": "NP25-100",
                    "Stain": "HE",
                    "BlockNumber": "B4",
                    "ParsingQCPassed": "TRUE",
                    "original_slide_path": "one.svs",
                },
                {
                    "AccessionID": "NP25-100",
                    "Stain": "HE",
                    "BlockNumber": "B4",
                    "ParsingQCPassed": "TRUE",
                    "original_slide_path": "two.svs",
                },
            ],
        )
        write_csv(
            self.batch / "completed_stages.csv",
            ["QC", "Renamed"],
            [{"QC": "True", "Renamed": "True"}],
        )
        renaming.atomic_write(
            self.batch / "name_mapping.csv",
            renaming.MAPPING_FIELDS,
            [
                mapping_row(
                    r"D:\scanner\misc\one.svs", "AAAAAA", "001"
                ),
                mapping_row(
                    r"D:\scanner\misc\two.svs", "AAAAAA", "002"
                ),
            ],
        )

        self.sdl_path = self.root / "Slide_Digitization_Log.xlsx"
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = app_module.Config.SDL_SHEET_NAME
        worksheet.append(app_module.SDL_HEADERS)
        row = [None] * len(app_module.SDL_HEADERS)
        row[app_module.SDL_HEADERS.index("Accession ID")] = "NP25-100"
        row[app_module.SDL_HEADERS.index("Date Loaded")] = app_module.datetime.date(
            2026, 7, 20
        )
        row[app_module.SDL_HEADERS.index("Pushed to SFTP Server")] = False
        worksheet.append(row)
        workbook.save(self.sdl_path)
        workbook.close()

        self.tq_home = self.root / ".tq"
        self.tq_home.mkdir()
        (self.tq_home / "config.toml").write_text(
            'username = "operator"\nftp_addr = "sftp.example"\nftp_dir = "/transfer"\n',
            encoding="utf-8",
        )
        self.original_config = {
            "LABEL_CHECK_BATCHES": app_module.Config.LABEL_CHECK_BATCHES,
            "INSTANCE_DIR": app_module.Config.INSTANCE_DIR,
            "SDL_FILE_PATH": app_module.Config.SDL_FILE_PATH,
            "TQ_HOME_DIR": app_module.Config.TQ_HOME_DIR,
            "TQ_TRANSFER_LOG_DIR": app_module.Config.TQ_TRANSFER_LOG_DIR,
            "TQ_EXECUTABLE": app_module.Config.TQ_EXECUTABLE,
            "IMAGE_STAGING_ROOT": app_module.Config.IMAGE_STAGING_ROOT,
            "IMAGE_STAGING_HOST_DISPLAY": (
                app_module.Config.IMAGE_STAGING_HOST_DISPLAY
            ),
        }
        app_module.Config.LABEL_CHECK_BATCHES = str(self.batch_base)
        app_module.Config.INSTANCE_DIR = str(self.root / "instance")
        app_module.Config.SDL_FILE_PATH = str(self.sdl_path)
        app_module.Config.TQ_HOME_DIR = str(self.tq_home)
        app_module.Config.TQ_TRANSFER_LOG_DIR = str(self.batch_base / "transfer_logs")
        app_module.Config.TQ_EXECUTABLE = "tq"
        app_module.Config.IMAGE_STAGING_ROOT = str(self.root / "image_staging")
        app_module.Config.IMAGE_STAGING_HOST_DISPLAY = r"D:\image_staging"
        app_module.batch_contexts.clear()
        with app_module._tq_state_lock:
            app_module._tq_drafts.clear()
            app_module._tq_jobs.clear()
            app_module._tq_active_job_id = None

        self.old_users = app_module.user_manager.users.copy()
        self.user = app_module.User("tq-user", "", False)
        self.admin = app_module.User("tq-admin", "", is_admin=True)
        app_module.user_manager.users[self.user.id] = self.user
        app_module.user_manager.users[self.admin.id] = self.admin
        app_module.app.config.update(TESTING=True, SECRET_KEY="tq-test")
        self.client = app_module.app.test_client()
        with self.client.session_transaction() as session:
            session["_user_id"] = self.user.id
            session["_fresh"] = True

    def login_as(self, user):
        with self.client.session_transaction() as session:
            session["_user_id"] = user.id
            session["_fresh"] = True

    def tearDown(self):
        for name, value in self.original_config.items():
            setattr(app_module.Config, name, value)
        app_module.batch_contexts.clear()
        with app_module._tq_state_lock:
            app_module._tq_drafts.clear()
            app_module._tq_jobs.clear()
            app_module._tq_active_job_id = None
        app_module.user_manager.users = self.old_users
        self.temporary.cleanup()

    def catalog(self):
        slides, warnings = app_module._tq_catalog()
        self.assertEqual([], warnings)
        self.assertEqual(2, len(slides))
        return slides

    def test_catalog_filters_dates_and_builds_destination(self):
        slides = self.catalog()

        self.assertEqual("2026-07-20", slides[0]["digitization_date"])
        self.assertEqual(
            [slides[0]],
            app_module._tq_filtered_slides(
                slides, "SectionCount", "001", "", "", "none"
            ),
        )
        self.assertEqual(
            "destination/BRAIN/AAAAAA",
            app_module._tq_destination_dir("destination", slides[0]),
        )
        with self.assertRaises(app_module.TQError):
            app_module._tq_destination_dir("../outside", slides[0])

    def test_page_filters_pid_and_contains_transfer_navigation(self):
        response = self.client.get("/tq?filter=PID&filter_value=AAAAAA")

        self.assertEqual(200, response.status_code)
        self.assertIn(b"Transfers", response.data)
        self.assertIn(b"Review Transfer", response.data)
        self.assertIn(b"AAAAAA", response.data)

    def test_transfer_console_uses_ansi_renderer(self):
        job = SimpleNamespace(
            id="ansi-job",
            owner_id=self.user.id,
            status="running",
            log_path=None,
        )
        with app_module._tq_state_lock:
            app_module._tq_jobs[job.id] = job
        with self.client.session_transaction() as session:
            session["tq_job_id"] = job.id

        response = self.client.get("/tq")

        self.assertEqual(200, response.status_code)
        self.assertIn(b"function appendAnsiOutput(text)", response.data)
        self.assertIn(b"applyAnsiCodes", response.data)
        self.assertIn(b"appendChild(span)", response.data)
        self.assertNotIn(
            b"consoleElement.textContent += data.output", response.data
        )

    def test_filter_validation_rejects_bad_typed_values_and_date_ranges(self):
        self.assertIn(
            "six uppercase letters",
            app_module._tq_validate_filter("PID", "ABC", "", ""),
        )
        self.assertIn(
            "three digits",
            app_module._tq_validate_filter("SectionCount", "12", "", ""),
        )
        self.assertIn(
            "cannot precede",
            app_module._tq_validate_filter(
                "AccessionDate", "", "2026-07-20", "2026-07-19"
            ),
        )

    def test_review_and_transfer_route_use_authoritative_mapping_values(self):
        slides = self.catalog()
        response = self.client.post(
            "/tq/review", data={"slide_id": slides[0]["id"]}
        )
        self.assertEqual(302, response.status_code)
        fake_job = SimpleNamespace(id="job-id", status="running")
        with mock.patch.object(
            app_module, "_start_tq_job", return_value=fake_job
        ) as start:
            response = self.client.post(
                "/tq/transfer",
                data={"destination_dir": "destination"},
            )

        self.assertEqual(302, response.status_code)
        launched = start.call_args.args[1]
        self.assertEqual("BRAIN", launched[0]["organ"])
        self.assertEqual("AAAAAA", launched[0]["pid"])
        self.assertEqual(
            "destination/BRAIN/AAAAAA", launched[0]["destination_dir"]
        )
        self.assertEqual("destination", launched[0]["staging_dir"])
        with self.client.session_transaction() as session:
            self.assertEqual("job-id", session["tq_job_id"])

    def test_config_requires_all_connection_values(self):
        (self.tq_home / "config.toml").write_text(
            'username = "operator"\nftp_addr = ""\nftp_dir = "/transfer"\n',
            encoding="utf-8",
        )
        with self.assertRaisesRegex(app_module.TQError, "ftp_addr"):
            app_module._tq_config()

    def run_result_job(self, slide, all_slides):
        slide = dict(slide)
        slide["destination_dir"] = app_module._tq_destination_dir(
            "destination", slide
        )
        output = (
            app_module.json.dumps(
                {
                    "original_path": slide["original_path"],
                    "success": True,
                    "error": None,
                }
            )
            + "\n"
        ).encode()
        manifest = self.root / f"{slide['id']}.csv"
        manifest.write_text("manifest\n", encoding="utf-8")
        job = app_module.TQJob(
            slide["id"],
            self.user.id,
            FakeProcess(output),
            [slide],
            all_slides,
            manifest,
        )
        with app_module._tq_state_lock:
            app_module._tq_jobs[job.id] = job
            app_module._tq_active_job_id = job.id
        app_module._read_tq_output(job)
        return job

    def test_logs_pid_and_updates_sdl_only_after_all_slides_succeed(self):
        slides = self.catalog()

        first_job = self.run_result_job(slides[0], slides)
        with first_job.log_path.open("r", newline="", encoding="utf-8") as handle:
            first_log = list(csv.DictReader(handle))
            self.assertEqual(list(app_module.TQ_LOG_FIELDS), list(first_log[0]))
            self.assertEqual("AAAAAA", first_log[0]["pid"])
            self.assertEqual("SUCCESS", first_log[0]["status"])
        workbook = load_workbook(self.sdl_path)
        worksheet = workbook[app_module.Config.SDL_SHEET_NAME]
        columns = app_module._sdl_header_columns(worksheet)
        self.assertFalse(
            worksheet.cell(
                row=2, column=columns["Pushed to SFTP Server"]
            ).value
        )
        workbook.close()

        second_job = self.run_result_job(slides[1], slides)

        self.assertEqual("succeeded", second_job.status)
        workbook = load_workbook(self.sdl_path)
        worksheet = workbook[app_module.Config.SDL_SHEET_NAME]
        columns = app_module._sdl_header_columns(worksheet)
        self.assertTrue(
            worksheet.cell(
                row=2, column=columns["Pushed to SFTP Server"]
            ).value
        )
        workbook.close()

    def test_launcher_schedules_staging_worker(self):
        slides = self.catalog()
        selected = [dict(slides[0])]
        selected[0]["destination_dir"] = app_module._tq_destination_dir(
            "destination", selected[0]
        )
        selected[0]["staging_dir"] = "destination"
        reader = mock.Mock()
        with mock.patch.object(
            app_module.threading, "Thread", return_value=reader
        ) as thread:
            job = app_module._start_tq_job(self.user.id, selected, slides)

        self.assertEqual(
            str(
                Path(app_module.Config.IMAGE_STAGING_ROOT)
                / "destination"
                / selected[0]["destination_name"]
            ),
            selected[0]["staged_path"],
        )
        self.assertEqual("running", job.status)
        self.assertIs(app_module._run_tq_job, thread.call_args.kwargs["target"])
        self.assertEqual((job,), thread.call_args.kwargs["args"])
        reader.start.assert_called_once_with()

    def test_worker_stages_deidentifies_then_uploads_and_cleans_success(self):
        slide = dict(self.catalog()[0])
        source = self.root / "gt450" / "source.svs"
        source.parent.mkdir()
        source.write_bytes(b"identifiable-slide")
        slide["original_path"] = str(source)
        slide["destination_dir"] = app_module._tq_destination_dir(
            "StudyA", slide
        )
        slide["staging_dir"] = "StudyA"
        app_module._tq_prepare_staging_paths([slide])
        job = app_module.TQJob(
            "stage-success", self.user.id, None, [slide], [slide]
        )
        captured = {}

        def launch(command, **kwargs):
            self.assertNotIn("shell", kwargs)
            self.assertNotIn("cwd", kwargs)
            if "deidentify_anonymize.py" in command[2]:
                staging_directory = Path(command[3])
                staged_path = str(next(staging_directory.glob("*.svs")))
                self.assertEqual(
                    b"identifiable-slide", Path(staged_path).read_bytes()
                )
                Path(staged_path).write_bytes(b"deidentified-slide")
                output_log = Path(command[command.index("--output-log") + 1])
                write_csv(
                    output_log,
                    ["original_path", "anonymized_path", "status"],
                    [
                        {
                            "original_path": staged_path,
                            "anonymized_path": staged_path,
                            "status": "SUCCESS",
                        }
                    ],
                )
                return FakeProcess(
                    b"Processing renamed slide\nAnonymization process completed.\n"
                )

            self.assertEqual(["tq", "pusher", "--paths"], command[:3])
            with Path(command[3]).open("r", newline="", encoding="utf-8") as handle:
                captured["manifest"] = list(csv.DictReader(handle))
            staged_path = captured["manifest"][0]["original_path"]
            self.assertEqual(b"deidentified-slide", Path(staged_path).read_bytes())
            output = (
                app_module.json.dumps(
                    {"original_path": staged_path, "success": True, "error": None}
                )
                + "\n"
            ).encode()
            return FakeProcess(output)

        with mock.patch.object(app_module.subprocess, "Popen", side_effect=launch):
            app_module._run_tq_job(job)

        manifest_row = captured["manifest"][0]
        self.assertEqual(slide["staged_path"], manifest_row["original_path"])
        self.assertEqual("StudyA/BRAIN/AAAAAA", manifest_row["destination_dir"])
        self.assertEqual("succeeded", job.status)
        self.assertFalse(Path(slide["staged_path"]).exists())
        output_markers = [
            "Downloaded D:\\image_staging",
            "Starting deidentify_anonymize.py",
            "Processing renamed slide",
            "deidentify_anonymize.py completed successfully",
            "Starting TQ upload",
        ]
        positions = [job.output.index(marker) for marker in output_markers]
        self.assertEqual(sorted(positions), positions)
        with job.log_path.open("r", newline="", encoding="utf-8") as handle:
            log_row = next(csv.DictReader(handle))
        self.assertEqual(str(source), log_row["original_path"])
        self.assertEqual("SUCCESS", log_row["status"])

    def test_deidentification_failure_prevents_every_upload_and_keeps_staging(self):
        slides = [dict(slide) for slide in self.catalog()]
        for index, slide in enumerate(slides, start=1):
            source = self.root / "gt450" / f"source-{index}.svs"
            source.parent.mkdir(exist_ok=True)
            source.write_bytes(f"slide-{index}".encode())
            slide["original_path"] = str(source)
            slide["destination_dir"] = app_module._tq_destination_dir(
                "StudyA", slide
            )
            slide["staging_dir"] = "StudyA"
        app_module._tq_prepare_staging_paths(slides)
        job = app_module.TQJob(
            "stage-failure", self.user.id, None, slides, slides
        )

        def fail_deidentification(command, **kwargs):
            self.assertIn("deidentify_anonymize.py", command[2])
            output_log = Path(command[command.index("--output-log") + 1])
            write_csv(
                output_log,
                ["original_path", "anonymized_path", "status"],
                [
                    {
                        "original_path": slides[1]["staged_path"],
                        "anonymized_path": slides[1]["staged_path"],
                        "status": "FAILURE",
                    }
                ],
            )
            return FakeProcess(b"Processing failed slide\n", return_code=1)

        with mock.patch.object(
            app_module.subprocess, "Popen", side_effect=fail_deidentification
        ) as popen:
            app_module._run_tq_job(job)

        self.assertEqual("failed", job.status)
        self.assertEqual(1, popen.call_count)
        self.assertTrue(Path(slides[0]["staged_path"]).exists())
        self.assertTrue(Path(slides[1]["staged_path"]).exists())
        self.assertIn("deidentify_anonymize.py exited with code 1", job.output)
        self.assertIsNotNone(job.log_path)

    def test_staging_rejects_duplicate_and_unsafe_destinations(self):
        slides = [dict(slide) for slide in self.catalog()]
        for slide in slides:
            slide["destination_dir"] = "StudyA/BRAIN/AAAAAA"
            slide["staging_dir"] = "StudyA"
            slide["destination_name"] = "same.svs"
        with self.assertRaisesRegex(app_module.TQError, "Duplicate"):
            app_module._tq_prepare_staging_paths(slides)

        slides[0]["staging_dir"] = "../outside"
        with self.assertRaisesRegex(app_module.TQError, "cannot be used"):
            app_module._tq_staging_path(slides[0])

    def test_log_browser_rejects_symlink_escape(self):
        outside = self.root / "outside.log"
        outside.write_text("secret", encoding="utf-8")
        (self.tq_home / "escape.log").symlink_to(outside)

        response = self.client.get("/tq/logs?file=escape.log")

        self.assertEqual(200, response.status_code)
        self.assertIn(b"Symbolic links cannot be opened", response.data)
        self.assertNotIn(b"secret", response.data)

    def test_log_browser_starts_with_folders_and_drills_into_files(self):
        log_file = self.tq_home / "logs" / "2026-07-24" / "transfer.log"
        log_file.parent.mkdir(parents=True)
        log_file.write_text("transfer complete", encoding="utf-8")
        loose_file = self.tq_home / "loose.log"
        loose_file.write_text("not shown at root", encoding="utf-8")

        root_response = self.client.get("/tq/logs")
        logs_response = self.client.get("/tq/logs?path=logs")
        date_response = self.client.get("/tq/logs?path=logs/2026-07-24")
        file_response = self.client.get(
            "/tq/logs?path=logs/2026-07-24&file=logs/2026-07-24/transfer.log"
        )

        self.assertEqual(200, root_response.status_code)
        self.assertIn(b"Log folders", root_response.data)
        self.assertIn(b"logs", root_response.data)
        self.assertNotIn(b"loose.log", root_response.data)
        self.assertIn(b"2026-07-24", logs_response.data)
        self.assertIn(b"transfer.log", date_response.data)
        self.assertIn(b"transfer complete", file_response.data)
        self.assertNotIn(b"Edit Config", root_response.data)

        self.login_as(self.admin)
        admin_response = self.client.get("/tq/logs")
        self.assertIn(b"Edit Config", admin_response.data)

    def test_config_editor_rejects_non_admin_without_changing_file(self):
        original = (self.tq_home / "config.toml").read_text(encoding="utf-8")

        get_response = self.client.get("/tq/config")
        post_response = self.client.post(
            "/tq/config", data={"config_text": 'username = "attacker"\n'}
        )

        self.assertEqual(403, get_response.status_code)
        self.assertEqual(403, post_response.status_code)
        self.assertEqual(
            original,
            (self.tq_home / "config.toml").read_text(encoding="utf-8"),
        )

    def test_config_editor_validates_and_atomically_saves_toml(self):
        self.login_as(self.admin)
        valid = (
            'username = "new-user"\n'
            'ftp_addr = "new.example"\n'
            'ftp_dir = "/new-transfer"\n'
        )

        response = self.client.post(
            "/tq/config", data={"config_text": valid}, follow_redirects=True
        )

        self.assertEqual(200, response.status_code)
        self.assertIn(b"saved successfully", response.data)
        self.assertEqual(
            valid,
            (self.tq_home / "config.toml").read_text(encoding="utf-8"),
        )

        invalid_response = self.client.post(
            "/tq/config", data={"config_text": 'username = "unterminated'}
        )

        self.assertEqual(200, invalid_response.status_code)
        self.assertIn(b"not valid TOML", invalid_response.data)
        self.assertEqual(
            valid,
            (self.tq_home / "config.toml").read_text(encoding="utf-8"),
        )


if __name__ == "__main__":
    unittest.main()
