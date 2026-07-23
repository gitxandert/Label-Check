"""
This script implements a Flask web application designed for the manual review and correction
of data extracted from whole-slide images (WSIs). It provides a user interface for operators
to verify and amend information like Accession IDs and Stain types that have been processed
by an automated pipeline (e.g., OCR).

The application features:
- User authentication (login/logout) with role-based access (standard user vs. admin).
- An admin panel for user management (adding new users).
- A robust queuing system that "leases" data rows to users for a fixed duration to prevent
  simultaneous edits. Expired leases are automatically returned to the queue.
- Dynamic loading and saving of data from/to a central CSV file.
- Automatic creation of backups before saving any changes.
- A user-friendly interface displaying slide images (macro/label) and form fields for data entry.
- Logic to pre-fill information based on other slides from the same patient/case.
- A command-line interface (CLI) for initializing the database and user accounts.
"""

# ==============================================================================
# 1. IMPORTS
# ==============================================================================
import csv
import codecs
import contextlib
import datetime
import functools
import hmac
import hashlib
import json
import logging
import os
import re
import secrets
import shutil
import sqlite3
import stat
import subprocess
import sys
import tempfile
import threading
import time
import uuid
from collections import Counter, defaultdict
from logging.handlers import RotatingFileHandler
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, Union

import click
import renaming

# Flask and its extensions for web framework, user management
from flask import (
    Flask,
    flash,
    get_flashed_messages,
    g,
    jsonify,
    redirect,
    render_template,
    request,
    send_from_directory,
    session,
    url_for,
)
from flask.cli import with_appcontext
from flask_login import (
    LoginManager,
    UserMixin,
    current_user,
    login_required,
    login_user,
    logout_user,
)
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.middleware.proxy_fix import ProxyFix
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

# ==============================================================================
# 2. CONFIGURATION
# ==============================================================================
class Config:
    """Central configuration class for the Flask application."""

    # A secret key is required for session management and security.
    SECRET_KEY = os.environ.get(
        "SECRET_KEY", "a-super-secret-key-that-you-should-change"
    )

    # --- Path Configuration ---
    # Robustly determine the project root directory.
    # We assume this file (app.py) is in <project_root>/src
    BASE_DIR = os.path.abspath(os.path.dirname(__file__))
    PROJECT_ROOT = os.path.dirname(BASE_DIR)

    # The base directory where all data (images, CSV) is located.
    # Using absolute path ensures we can run the app from anywhere.
    IMAGE_BASE_DIR = PROJECT_ROOT

    # The full path to the primary CSV file.
    CSV_FILE_PATH = os.path.join(IMAGE_BASE_DIR, "enriched.csv")
    
    # Directory to store timestamped backups.
    BACKUP_DIR = os.path.join(BASE_DIR, "csv_backups")

    # Instance directory for local data persistence
    INSTANCE_DIR = os.path.join(BASE_DIR, "instance")
    
    # CSV persistence files
    USERS_CSV_PATH = os.path.join(INSTANCE_DIR, "users.csv")
    QUEUE_CSV_PATH = os.path.join(INSTANCE_DIR, "queue.csv")
    API_DB_PATH = os.path.join(INSTANCE_DIR, "api.sqlite3")
    API_JOB_OUTPUT_DIR = os.path.join(INSTANCE_DIR, "pipeline_job_output")
    API_REQUIRE_HTTPS = os.environ.get("API_REQUIRE_HTTPS", "true").lower() == "true"
    API_TRUST_PROXY_HEADERS = os.environ.get("API_TRUST_PROXY_HEADERS", "false").lower() == "true"
    API_SUBMIT_RATE_LIMIT = int(os.environ.get("API_SUBMIT_RATE_LIMIT", "5"))
    API_READ_RATE_LIMIT = int(os.environ.get("API_READ_RATE_LIMIT", "60"))
    API_RATE_WINDOW_SECONDS = 60
    API_DEFAULT_TOKEN_DAYS = 90
    API_OUTPUT_DEFAULT_LIMIT = 16 * 1024
    API_OUTPUT_MAX_LIMIT = 64 * 1024
    CSRF_ENABLED = os.environ.get("CSRF_ENABLED", "true").lower() == "true"

    # Slide Digitization Log workbook configuration.
    SDL_FILE_PATH = os.path.join(BASE_DIR, "logs", "Slide_Digitization_Log.xlsx")
    SDL_SHEET_NAME = "general"
    SDL_ORGANS = ("BRAIN", "BREAST", "TESTES", "CYTO")
    SDL_SCANNERS = ("-----", "RSCH1 (SS12797)", "CLIN1 (SS12602)")


    # Path to scanner inventories
    SCANNER_INVENTORIES = "D:\\scanner_inventories"
    # Path to batches of new slides to label-check
    LABEL_CHECK_BATCHES = "D:\\label_check_batches"
    COPATH_CLONE = os.environ.get("COPATH_CLONE", "D:\\copath_clone")

    # Default password for the initial 'admin' user.
    ADMIN_DEFAULT_PASSWORD = os.environ.get(
        "ADMIN_DEFAULT_PASSWORD", "change_this_password"
    )

    # --- Queue Settings ---
    # The duration (in seconds) a user can hold a "lease" on a queue item before it's
    # automatically returned to the pool for others.
    LEASE_DURATION_SECONDS = 300  # 5 minutes


# ==============================================================================
# 3. LOGGING SETUP
# ==============================================================================
def setup_logging(app: Flask) -> None:
    """Configures comprehensive logging for the application."""
    if not os.path.exists("logs"):
        os.mkdir("logs")

    file_handler = RotatingFileHandler("logs/app.log", maxBytes=102400, backupCount=10)
    file_handler.setFormatter(
        logging.Formatter(
            "%(asctime)s %(levelname)s: %(message)s [in %(pathname)s:%(lineno)d]"
        )
    )
    file_handler.setLevel(logging.INFO)

    console_handler = logging.StreamHandler()
    console_handler.setFormatter(
        logging.Formatter("%(asctime)s - %(levelname)s - %(message)s")
    )
    console_handler.setLevel(logging.INFO)

    app.logger.addHandler(file_handler)
    app.logger.addHandler(console_handler)
    app.logger.setLevel(logging.INFO)
    app.logger.info("Application startup")


# ==============================================================================
# 4. APPLICATION & EXTENSIONS INITIALIZATION
# ==============================================================================
base_dir = os.path.abspath(os.path.dirname(__file__))
instance_path = os.path.join(base_dir, "instance")
template_dir = os.path.join(base_dir, "templates")

app = Flask(__name__, template_folder=template_dir, instance_path=instance_path)
app.config.from_object(Config)
if app.config["API_TRUST_PROXY_HEADERS"]:
    app.wsgi_app = ProxyFix(app.wsgi_app, x_proto=1)
os.makedirs(app.instance_path, exist_ok=True)

setup_logging(app)

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = "login"
login_manager.login_message = "Please log in to access this page."
login_manager.login_message_category = "warning"


# ==============================================================================
# 5. CUSTOM EXCEPTIONS
# ==============================================================================
class DataLoadError(Exception):
    pass

class DataSaveError(Exception):
    pass

class BackupError(Exception):
    pass

class SDLWorkbookError(Exception):
    pass

class SDLValidationError(Exception):
    pass


class InventoryReadError(Exception):
    pass


def _utcnow() -> datetime.datetime:
    return datetime.datetime.now(datetime.timezone.utc)


def _iso_utc(value: Optional[datetime.datetime] = None) -> str:
    return (value or _utcnow()).isoformat().replace("+00:00", "Z")


class APIStore:
    """SQLite-backed API credentials, durable job metadata, and rate counters."""

    def __init__(self, db_path: str, output_dir: str):
        self.db_path = db_path
        self.output_dir = output_dir
        self._initialized_path: Optional[str] = None
        self._init_lock = threading.Lock()

    def configure(self, db_path: str, output_dir: str) -> None:
        self.db_path = db_path
        self.output_dir = output_dir
        self._initialized_path = None

    def _connect(self) -> sqlite3.Connection:
        self._ensure_schema()
        connection = sqlite3.connect(self.db_path, timeout=30)
        connection.row_factory = sqlite3.Row
        return connection

    @contextlib.contextmanager
    def connection(self):
        connection = self._connect()
        try:
            with connection:
                yield connection
        finally:
            connection.close()

    def _ensure_schema(self) -> None:
        if self._initialized_path == self.db_path:
            return
        with self._init_lock:
            if self._initialized_path == self.db_path:
                return
            Path(self.db_path).parent.mkdir(parents=True, exist_ok=True)
            Path(self.output_dir).mkdir(parents=True, exist_ok=True)
            connection = sqlite3.connect(self.db_path, timeout=30)
            try:
                connection.executescript(
                    """
                    CREATE TABLE IF NOT EXISTS api_tokens (
                        token_id TEXT PRIMARY KEY,
                        user_id TEXT NOT NULL,
                        label TEXT NOT NULL,
                        secret_hash TEXT NOT NULL,
                        scopes TEXT NOT NULL,
                        created_at TEXT NOT NULL,
                        expires_at TEXT NOT NULL,
                        revoked_at TEXT,
                        last_used_at TEXT
                    );
                    CREATE TABLE IF NOT EXISTS pipeline_jobs (
                        job_id TEXT PRIMARY KEY,
                        owner_id TEXT NOT NULL,
                        token_id TEXT,
                        idempotency_key TEXT,
                        payload_hash TEXT,
                        request_json TEXT NOT NULL,
                        command_json TEXT NOT NULL,
                        status TEXT NOT NULL,
                        created_at TEXT NOT NULL,
                        started_at TEXT,
                        completed_at TEXT,
                        return_code INTEGER,
                        output_path TEXT NOT NULL,
                        launcher_pid INTEGER,
                        UNIQUE(token_id, idempotency_key)
                    );
                    CREATE UNIQUE INDEX IF NOT EXISTS one_active_pipeline_job
                    ON pipeline_jobs((1)) WHERE status IN ('starting', 'running');
                    CREATE TABLE IF NOT EXISTS api_rate_limits (
                        token_id TEXT NOT NULL,
                        bucket TEXT NOT NULL,
                        window_start INTEGER NOT NULL,
                        request_count INTEGER NOT NULL,
                        PRIMARY KEY(token_id, bucket)
                    );
                    """
                )
                connection.commit()
            finally:
                connection.close()
            self._initialized_path = self.db_path

    def create_token(
        self, user_id: str, label: str, scopes: List[str], expires_days: int
    ) -> Tuple[str, Dict[str, Any]]:
        token_id = uuid.uuid4().hex[:16]
        raw_token = f"lc_pat_{token_id}.{secrets.token_urlsafe(32)}"
        created = _utcnow()
        expires = created + datetime.timedelta(days=expires_days)
        record = {
            "token_id": token_id,
            "user_id": user_id,
            "label": label,
            "scopes": sorted(set(scopes)),
            "created_at": _iso_utc(created),
            "expires_at": _iso_utc(expires),
            "revoked_at": None,
            "last_used_at": None,
        }
        with self.connection() as connection:
            connection.execute(
                """INSERT INTO api_tokens
                   (token_id, user_id, label, secret_hash, scopes, created_at, expires_at)
                   VALUES (?, ?, ?, ?, ?, ?, ?)""",
                (
                    token_id,
                    user_id,
                    label,
                    hashlib.sha256(raw_token.encode("utf-8")).hexdigest(),
                    " ".join(record["scopes"]),
                    record["created_at"],
                    record["expires_at"],
                ),
            )
        return raw_token, record

    def list_tokens(self, user_id: Optional[str] = None) -> List[Dict[str, Any]]:
        query = "SELECT * FROM api_tokens"
        params: Tuple[Any, ...] = ()
        if user_id:
            query += " WHERE user_id = ?"
            params = (user_id,)
        query += " ORDER BY created_at DESC"
        with self.connection() as connection:
            rows = connection.execute(query, params).fetchall()
        return [self._token_record(row) for row in rows]

    @staticmethod
    def _token_record(row: sqlite3.Row) -> Dict[str, Any]:
        return {
            "token_id": row["token_id"],
            "user_id": row["user_id"],
            "label": row["label"],
            "scopes": row["scopes"].split(),
            "created_at": row["created_at"],
            "expires_at": row["expires_at"],
            "revoked_at": row["revoked_at"],
            "last_used_at": row["last_used_at"],
        }

    def authenticate_token(self, raw_token: str) -> Optional[Dict[str, Any]]:
        match = re.fullmatch(r"lc_pat_([0-9a-f]{16})\.[A-Za-z0-9_-]+", raw_token)
        if not match:
            return None
        with self.connection() as connection:
            row = connection.execute(
                "SELECT * FROM api_tokens WHERE token_id = ?", (match.group(1),)
            ).fetchone()
            if row is None or row["revoked_at"]:
                return None
            expected = row["secret_hash"]
            actual = hashlib.sha256(raw_token.encode("utf-8")).hexdigest()
            if not hmac.compare_digest(expected, actual):
                return None
            expires = datetime.datetime.fromisoformat(row["expires_at"].replace("Z", "+00:00"))
            if expires <= _utcnow():
                return None
            used_at = _iso_utc()
            connection.execute(
                "UPDATE api_tokens SET last_used_at = ? WHERE token_id = ?",
                (used_at, row["token_id"]),
            )
            record = self._token_record(row)
            record["last_used_at"] = used_at
            return record

    def revoke_token(self, token_id: str) -> bool:
        with self.connection() as connection:
            cursor = connection.execute(
                "UPDATE api_tokens SET revoked_at = ? WHERE token_id = ? AND revoked_at IS NULL",
                (_iso_utc(), token_id),
            )
            return cursor.rowcount == 1

    def rate_limit(self, token_id: str, bucket: str, limit: int, window: int) -> Tuple[bool, int, int]:
        now = int(time.time())
        window_start = now - (now % window)
        with self.connection() as connection:
            connection.execute("BEGIN IMMEDIATE")
            row = connection.execute(
                "SELECT window_start, request_count FROM api_rate_limits WHERE token_id=? AND bucket=?",
                (token_id, bucket),
            ).fetchone()
            count = 0 if row is None or row["window_start"] != window_start else row["request_count"]
            allowed = count < limit
            if allowed:
                count += 1
                connection.execute(
                    """INSERT INTO api_rate_limits(token_id, bucket, window_start, request_count)
                       VALUES (?, ?, ?, ?)
                       ON CONFLICT(token_id, bucket) DO UPDATE SET
                       window_start=excluded.window_start, request_count=excluded.request_count""",
                    (token_id, bucket, window_start, count),
                )
        return allowed, max(0, limit - count), window_start + window - now

    def find_idempotent(self, token_id: str, key: str) -> Optional[sqlite3.Row]:
        with self.connection() as connection:
            return connection.execute(
                "SELECT * FROM pipeline_jobs WHERE token_id=? AND idempotency_key=?",
                (token_id, key),
            ).fetchone()

    def reserve_job(
        self,
        job_id: str,
        owner_id: str,
        values: Dict[str, Any],
        command: List[str],
        token_id: Optional[str] = None,
        idempotency_key: Optional[str] = None,
        payload_hash: Optional[str] = None,
    ) -> str:
        self._ensure_schema()
        output_path = str(Path(self.output_dir) / f"{job_id}.log")
        Path(output_path).touch(exist_ok=False)
        try:
            with self.connection() as connection:
                connection.execute(
                    """INSERT INTO pipeline_jobs
                       (job_id, owner_id, token_id, idempotency_key, payload_hash,
                        request_json, command_json, status, created_at, output_path)
                       VALUES (?, ?, ?, ?, ?, ?, ?, 'starting', ?, ?)""",
                    (
                        job_id, owner_id, token_id, idempotency_key, payload_hash,
                        json.dumps(values, sort_keys=True), json.dumps(command), _iso_utc(), output_path,
                    ),
                )
        except Exception:
            os.remove(output_path)
            raise
        return output_path

    def update_job(self, job_id: str, **fields: Any) -> None:
        allowed = {"status", "started_at", "completed_at", "return_code", "launcher_pid"}
        selected = {key: value for key, value in fields.items() if key in allowed}
        if not selected:
            return
        assignments = ", ".join(f"{key}=?" for key in selected)
        with self.connection() as connection:
            connection.execute(
                f"UPDATE pipeline_jobs SET {assignments} WHERE job_id=?",
                (*selected.values(), job_id),
            )

    def get_job(self, job_id: str) -> Optional[Dict[str, Any]]:
        with self.connection() as connection:
            row = connection.execute(
                "SELECT * FROM pipeline_jobs WHERE job_id=?", (job_id,)
            ).fetchone()
        return dict(row) if row else None

    def mark_stale_jobs_interrupted(self) -> None:
        with self.connection() as connection:
            rows = connection.execute(
                "SELECT job_id, launcher_pid FROM pipeline_jobs WHERE status IN ('starting', 'running')"
            ).fetchall()
            for row in rows:
                launcher_pid = row["launcher_pid"]
                if launcher_pid:
                    try:
                        os.kill(launcher_pid, 0)
                        continue
                    except (OSError, ProcessLookupError):
                        pass
                connection.execute(
                    """UPDATE pipeline_jobs SET status='interrupted', completed_at=?
                       WHERE job_id=? AND status IN ('starting', 'running')""",
                    (_iso_utc(), row["job_id"]),
                )


api_store = APIStore(Config.API_DB_PATH, Config.API_JOB_OUTPUT_DIR)
api_store.mark_stale_jobs_interrupted()


# ==============================================================================
# 6. PERSISTENCE MODELS (CSV BASED)
# ==============================================================================
class User(UserMixin):
    """Represents a user account."""
    def __init__(self, id: str, password_hash: str, correction_count: int = 0, is_admin: bool = False):
        self.id = id
        self.password_hash = password_hash
        self.correction_count = int(correction_count)
        # Handle string 'True'/'False' from CSV loading
        if isinstance(is_admin, str):
            self.is_admin = is_admin.lower() == 'true'
        else:
            self.is_admin = bool(is_admin)

    def set_password(self, password: str) -> None:
        self.password_hash = generate_password_hash(password)

    def verify_password(self, password: str) -> bool:
        return check_password_hash(self.password_hash, password)

    def to_dict(self) -> Dict[str, str]:
        return {
            "id": self.id,
            "password_hash": self.password_hash,
            "correction_count": str(self.correction_count),
            "is_admin": str(self.is_admin)
        }

    def __repr__(self) -> str:
        return f"<User {self.id}>"


class QueueItem:
    """Represents a single row from the CSV in the processing queue."""
    def __init__(self, original_index: int, status: str = "pending", 
                 leased_by_id: Optional[str] = None, leased_at: Optional[Union[str, datetime.datetime]] = None,
                 completed_by_id: Optional[str] = None, completed_at: Optional[Union[str, datetime.datetime]] = None,
                 row_id: Optional[int] = None):
        self.id = row_id # ID is strictly internal/optional for QueueItem in this CSV context, but we keep track if needed.
        self.original_index = int(original_index)
        self.status = status
        self.leased_by_id = leased_by_id if leased_by_id != "" else None
        
        # Date parsing logic
        self.leased_at = self._parse_date(leased_at)
        self.completed_by_id = completed_by_id if completed_by_id != "" else None
        self.completed_at = self._parse_date(completed_at)

    def _parse_date(self, date_val: Union[str, datetime.datetime, None]) -> Optional[datetime.datetime]:
        if not date_val:
            return None
        if isinstance(date_val, datetime.datetime):
            return date_val
        try:
            return datetime.datetime.fromisoformat(date_val)
        except ValueError:
            return None

    def _format_date(self, date_val: Optional[datetime.datetime]) -> str:
        return date_val.isoformat() if date_val else ""

    def to_dict(self) -> Dict[str, str]:
        return {
            "original_index": str(self.original_index),
            "status": self.status,
            "leased_by_id": self.leased_by_id if self.leased_by_id else "",
            "leased_at": self._format_date(self.leased_at),
            "completed_by_id": self.completed_by_id if self.completed_by_id else "",
            "completed_at": self._format_date(self.completed_at)
        }

    @property
    def leased_by(self):
        """Helper to resolve user object for template compatibility."""
        if self.leased_by_id:
            # Access global user_manager
            return user_manager.get(self.leased_by_id)
        return None

    def __repr__(self) -> str:
        return f"<QueueItem {self.original_index} - {self.status}>"


class CSVManager:
    """Generic CSV persistence manager."""
    def __init__(self, filepath: str, fieldnames: List[str]):
        self.filepath = filepath
        self.fieldnames = fieldnames
        self._lock = threading.Lock()

    def _ensure_file(self):
        if not os.path.exists(self.filepath):
            with open(self.filepath, 'w', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=self.fieldnames)
                writer.writeheader()

    def read_all(self) -> List[Dict[str, str]]:
        self._ensure_file()
        with self._lock:
            try:
                with open(self.filepath, 'r', newline='', encoding='utf-8') as f:
                    reader = csv.DictReader(f)
                    return list(reader)
            except Exception as e:
                app.logger.error(f"Error reading {self.filepath}: {e}")
                return []

    def write_all(self, data: List[Dict[str, str]]) -> None:
        with self._lock:
            try:
                # write atomic
                temp_path = self.filepath + ".tmp"
                with open(temp_path, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.DictWriter(f, fieldnames=self.fieldnames)
                    writer.writeheader()
                    writer.writerows(data)
                
                os.replace(temp_path, self.filepath)
            except Exception as e:
                app.logger.error(f"Error writing {self.filepath}: {e}")
                if os.path.exists(temp_path):
                    os.remove(temp_path)
                raise DataSaveError(f"Failed to save CSV {self.filepath}: {e}")


class UserManager(CSVManager):
    def __init__(self):
        super().__init__(Config.USERS_CSV_PATH, ["id", "password_hash", "correction_count", "is_admin"])
        # Cache users in memory for performance, similar to DB
        self.users: Dict[str, User] = {}
        self.load()

    def load(self):
        rows = self.read_all()
        self.users = {}
        for row in rows:
            u = User(
                id=row["id"],
                password_hash=row["password_hash"],
                correction_count=int(row["correction_count"]),
                is_admin=row["is_admin"]
            )
            self.users[u.id] = u

    def save(self):
        data = [u.to_dict() for u in self.users.values()]
        self.write_all(data)

    def get(self, user_id: str) -> Optional[User]:
        return self.users.get(user_id)

    def add(self, user: User):
        self.users[user.id] = user
        self.save()  # Auto-save on add due to simple architecture

    def update(self, user: User):
        self.users[user.id] = user
        self.save()
    
    def get_all(self) -> List[User]:
        return list(self.users.values())


class QueueManager(CSVManager):
    def __init__(self, filepath: str = Config.QUEUE_CSV_PATH):
        super().__init__(filepath, ["original_index", "status", "leased_by_id", "leased_at", "completed_by_id", "completed_at"])
        self.items: Dict[int, QueueItem] = {}
        self.load()

    def load(self):
        rows = self.read_all()
        self.items = {}
        for row in rows:
            try:
                idx = int(row["original_index"])
                item = QueueItem(
                    original_index=idx,
                    status=row["status"],
                    leased_by_id=row.get("leased_by_id"),
                    leased_at=row.get("leased_at"),
                    completed_by_id=row.get("completed_by_id"),
                    completed_at=row.get("completed_at"),
                )
                self.items[idx] = item
            except ValueError:
                continue

    def save(self):
        data = [item.to_dict() for item in self.items.values()]
        self.write_all(data)

    def get(self, original_index: int) -> Optional[QueueItem]:
        return self.items.get(original_index)
    
    def add(self, item: QueueItem):
        self.items[item.original_index] = item
        # Batch add usually calls save manually, but for single integrity:
        # self.save() 
    
    def get_all(self) -> List[QueueItem]:
        return list(self.items.values())

    def update(self):
        """Persist current state."""
        self.save()


# Initialize Managers
user_manager = UserManager()


@login_manager.user_loader
def load_user(user_id: str) -> Optional[User]:
    return user_manager.get(user_id)


# ==============================================================================
# 7. DATA MANAGER
# ==============================================================================
class DataManager:
    """Manages the in-memory CSV data state, loading, and saving."""
    def __init__(self, batch_root: Optional[Path] = None, csv_path: Optional[Path] = None):
        self.data: List[Dict[str, Any]] = []
        self.headers: List[str] = []
        self.batch_root = batch_root
        self.csv_path = csv_path
        self._lock = threading.Lock() # Ensure thread safety for data access
        self.critical_headers = ["AccessionID", "Stain", "ParsingQCPassed", "original_slide_path"]

    def load_data(self, file_path: Optional[Union[str, Path]] = None) -> None:
        """Loads CSV data into memory safely."""
        file_path = str(file_path or self.csv_path or Config.CSV_FILE_PATH)
        with self._lock:
            app.logger.info(f"Loading CSV data from: {file_path}")
            if not os.path.exists(file_path):
                raise DataLoadError(f"CSV file not found: {file_path}")

            _data: List[Dict[str, Any]] = []
            try:
                with open(file_path, "r", newline="", encoding="utf-8") as csvfile:
                    reader = csv.DictReader(csvfile, delimiter=",")
                    _headers = reader.fieldnames

                    if not _headers:
                        raise DataLoadError("CSV file is empty or has no header.")
                    
                    missing = [h for h in self.critical_headers if h not in _headers]
                    if missing:
                        app.logger.warning(
                            f"CSV is missing expected headers: {missing}. Functionality may be limited."
                        )

                    for i, row in enumerate(reader):
                        row["_original_index"] = i
                        orig_path = row.get("original_slide_location")
                        row["_identifier"] = Path(orig_path).stem if orig_path else f"Unknown_{i}"
                        row["_label_text"] = row.get("label_text", "N/A")
                        row["_macro_text"] = row.get("macro_text", "N/A")
                        row["_label_path"] = row.get("label_path")
                        row["_macro_path"] = row.get("macro_path")
                        
                        row["AccessionID"] = row.get("AccessionID", "").strip()
                        row["Stain"] = row.get("Stain", "").strip()
                        row["BlockNumber"] = row.get("BlockNumber", "").strip()
                        
                        qc_passed_str = row.get("ParsingQCPassed", "").strip()
                        row["_is_complete"] = bool(
                            qc_passed_str and qc_passed_str.lower() != "false"
                        )
                        _data.append(row)

                # Post-processing: Calculate per-patient file statistics
                patient_slide_ids = defaultdict(list)
                for i, row in enumerate(_data):
                    patient_slide_ids[row["_identifier"]].append(i)
                
                for _, original_indices in patient_slide_ids.items():
                    total = len(original_indices)
                    for j, original_idx in enumerate(sorted(original_indices)):
                        _data[original_idx]["_total_patient_files"] = total
                        _data[original_idx]["_patient_file_number"] = j + 1

                self.data = _data
                self.headers = _headers
                self._recalculate_accession_counts()
                app.logger.info(f"Loaded {len(self.data)} rows.")

            except Exception as e:
                self.data, self.headers = [], []
                raise DataLoadError(f"Error reading CSV: {e}")

    def save_data(self, target_path: Optional[Union[str, Path]] = None) -> None:
        """Saves current data to CSV atomically."""
        target_path = str(target_path or self.csv_path or Config.CSV_FILE_PATH)
        with self._lock:
            if not self.data or not self.headers:
                app.logger.warning("Save aborted: No data in memory.")
                return

            app.logger.info(f"Saving {len(self.data)} rows to {target_path}")

            priority_fields = ["AccessionID", "Stain", "BlockNumber", "ParsingQCPassed"]
            pipeline_fields = [h for h in self.headers if h not in priority_fields]
            fieldnames = list(dict.fromkeys(priority_fields + pipeline_fields))
            
            temp_path = target_path + ".tmp"
            try:
                with open(temp_path, "w", newline="", encoding="utf-8") as csvfile:
                    writer = csv.DictWriter(
                        csvfile,
                        fieldnames=fieldnames,
                        delimiter=",",
                        extrasaction="ignore",
                        quoting=csv.QUOTE_MINIMAL,
                    )
                    writer.writeheader()

                    for row in self.data:
                        write_row = row.copy()
                        write_row["ParsingQCPassed"] = "TRUE" if row.get("_is_complete") else ""
                        writer.writerow(write_row)

                # Atomic replace
                if os.path.exists(target_path):
                    os.replace(temp_path, target_path)
                else:
                    os.rename(temp_path, target_path)
                    
                session["last_loaded_csv_mod_time"] = os.path.getmtime(target_path)
                app.logger.info("Save successful.")
            except Exception as e:
                if os.path.exists(temp_path):
                    os.remove(temp_path)
                raise DataSaveError(f"Failed to save CSV: {e}")

    def _recalculate_accession_counts(self) -> None:
        """Internal helper to count AccessionID occurrences."""
        if not self.data:
            return
        id_counts = Counter(
            row.get("AccessionID", "").strip()
            for row in self.data
            if row.get("AccessionID", "").strip()
        )
        for row in self.data:
            current_id = row.get("AccessionID", "").strip()
            row["_accession_id_count"] = id_counts[current_id] if current_id else 0
    
    def get_row(self, index: int) -> Optional[Dict[str, Any]]:
        with self._lock:
            if 0 <= index < len(self.data):
                return self.data[index]
            return None

    def update_row(self, index: int, updates: Dict[str, Any]) -> bool:
        """Updates a row and triggers recalculations if needed."""
        with self._lock:
            if not (0 <= index < len(self.data)):
                return False
            
            row = self.data[index]
            has_changed = False
            recalc_counts = False

            for key, value in updates.items():
                if row.get(key) != value:
                    row[key] = value
                    has_changed = True
                    if key == "AccessionID":
                        recalc_counts = True
            
            if recalc_counts:
                self._recalculate_accession_counts()
                
            return has_changed

    def clear(self):
        with self._lock:
            self.data = []
            self.headers = []

    def get_absolute_path(self, relative_path: str) -> Optional[str]:
        """Resolve a CSV image path, constrained to the active batch directory."""
        if not relative_path:
            return None

        root = (self.batch_root or Path(Config.IMAGE_BASE_DIR)).resolve()
        cleaned_path = str(relative_path).replace("\\", os.sep)
        candidate = Path(cleaned_path)
        if not candidate.is_absolute():
            candidate = root / candidate
        try:
            resolved = candidate.resolve()
            if os.path.commonpath([str(root), str(resolved)]) != str(root):
                return None
        except (OSError, ValueError):
            return None
        return str(resolved)

    def check_paths(self) -> List[str]:
        """Checks if all image paths in the loaded data exist and are readable."""
        missing_or_unreadable = []
        with self._lock:
            for i, row in enumerate(self.data):
                # Check Label and Macro paths
                for key in ["_label_path", "_macro_path"]:
                    rel_path = row.get(key)
                    if rel_path:
                        abs_path = self.get_absolute_path(rel_path)
                        if not abs_path or not os.path.exists(abs_path):
                             missing_or_unreadable.append(f"Row {i+1} ({key}): Path not found -> {abs_path}")
                        elif not os.access(abs_path, os.R_OK):
                             missing_or_unreadable.append(f"Row {i+1} ({key}): Path not readable -> {abs_path}")
        return missing_or_unreadable

class BatchContext:
    """Loaded data and persistent queue belonging to one discovered batch."""

    COMPLETED_STAGES_FIELDS = ["QC", "Renamed"]

    def __init__(self, batch_id: str, root: Path):
        self.id = batch_id
        self.root = root
        self.name = root.name
        self.display_name = f"{root.parent.name}/{root.name}"
        self.csv_path = root / "enriched.csv"
        self.completed_stages_path = root / "completed_stages.csv"
        queue_dir = Path(Config.INSTANCE_DIR) / "batch_queues"
        queue_dir.mkdir(parents=True, exist_ok=True)
        self.data_manager = DataManager(root, self.csv_path)
        self.queue_manager = QueueManager(str(queue_dir / f"{batch_id}.csv"))
        self.csv_mod_time: Optional[float] = None
        self.completed_stages = {"QC": False, "Renamed": False}
        self._completed_stages_lock = threading.Lock()

    @staticmethod
    def _parse_stage_value(value: Optional[str], column: str) -> bool:
        normalized = (value or "").strip().lower()
        if normalized not in {"true", "false"}:
            raise DataLoadError(
                f"completed_stages.csv has an invalid {column} value"
            )
        return normalized == "true"

    def _write_new_completed_stages(self) -> None:
        """Atomically create the default stage file without replacing an existing one."""
        temp_path = self.completed_stages_path.with_name(
            f".{self.completed_stages_path.name}.{uuid.uuid4().hex}.tmp"
        )
        try:
            with open(temp_path, "x", newline="", encoding="utf-8") as csvfile:
                writer = csv.DictWriter(
                    csvfile, fieldnames=self.COMPLETED_STAGES_FIELDS
                )
                writer.writeheader()
                writer.writerow({"QC": "False", "Renamed": "False"})
                csvfile.flush()
                os.fsync(csvfile.fileno())
            try:
                os.link(temp_path, self.completed_stages_path)
            except FileExistsError:
                # Another discovery request initialized the batch first.
                pass
        finally:
            try:
                temp_path.unlink()
            except FileNotFoundError:
                pass

    def load_completed_stages(self, create_if_missing: bool = False) -> None:
        """Load the single-row batch stage file, optionally initializing it."""
        with self._completed_stages_lock:
            if create_if_missing and not self.completed_stages_path.exists():
                self._write_new_completed_stages()
            try:
                with open(
                    self.completed_stages_path,
                    "r",
                    newline="",
                    encoding="utf-8",
                ) as csvfile:
                    reader = csv.DictReader(csvfile)
                    if reader.fieldnames != self.COMPLETED_STAGES_FIELDS:
                        raise DataLoadError(
                            "completed_stages.csv must have QC and Renamed columns"
                        )
                    rows = list(reader)
            except DataLoadError:
                raise
            except (OSError, UnicodeError, csv.Error) as exc:
                raise DataLoadError(
                    f"could not read completed_stages.csv: {exc}"
                ) from exc

            if len(rows) != 1:
                raise DataLoadError(
                    "completed_stages.csv must contain exactly one data row"
                )
            row = rows[0]
            self.completed_stages = {
                column: self._parse_stage_value(row.get(column), column)
                for column in self.COMPLETED_STAGES_FIELDS
            }

    def mark_qc_complete(self) -> None:
        """Atomically mark the batch QC stage complete and preserve Renamed."""
        with self._completed_stages_lock:
            renamed = self.completed_stages["Renamed"]
            temp_path = self.completed_stages_path.with_name(
                f".{self.completed_stages_path.name}.{uuid.uuid4().hex}.tmp"
            )
            try:
                with open(temp_path, "x", newline="", encoding="utf-8") as csvfile:
                    writer = csv.DictWriter(
                        csvfile, fieldnames=self.COMPLETED_STAGES_FIELDS
                    )
                    writer.writeheader()
                    writer.writerow(
                        {
                            "QC": "True",
                            "Renamed": "True" if renamed else "False",
                        }
                    )
                    csvfile.flush()
                    os.fsync(csvfile.fileno())
                os.replace(temp_path, self.completed_stages_path)
                self.completed_stages["QC"] = True
            except OSError as exc:
                try:
                    temp_path.unlink()
                except FileNotFoundError:
                    pass
                raise DataSaveError(
                    f"could not update completed_stages.csv: {exc}"
                ) from exc

    def mark_renamed_complete(self) -> None:
        """Atomically mark renaming complete while preserving QC."""
        with self._completed_stages_lock:
            temp_path = self.completed_stages_path.with_name(
                f".{self.completed_stages_path.name}.{uuid.uuid4().hex}.tmp"
            )
            try:
                with open(temp_path, "x", newline="", encoding="utf-8") as csvfile:
                    writer = csv.DictWriter(csvfile, fieldnames=self.COMPLETED_STAGES_FIELDS)
                    writer.writeheader()
                    writer.writerow({"QC": "True", "Renamed": "True"})
                    csvfile.flush()
                    os.fsync(csvfile.fileno())
                os.replace(temp_path, self.completed_stages_path)
                self.completed_stages = {"QC": True, "Renamed": True}
            except OSError as exc:
                temp_path.unlink(missing_ok=True)
                raise DataSaveError(f"could not update completed_stages.csv: {exc}") from exc

    @property
    def qc_complete(self) -> bool:
        return self.completed_stages["QC"]

    def refresh(self) -> None:
        mod_time = self.csv_path.stat().st_mtime
        if not self.data_manager.data or mod_time != self.csv_mod_time:
            self.data_manager.load_data(self.csv_path)
            self.csv_mod_time = mod_time
        valid_indices = set(range(len(self.data_manager.data)))
        changed = False
        for index in list(self.queue_manager.items):
            if index not in valid_indices:
                del self.queue_manager.items[index]
                changed = True
        for row in self.data_manager.data:
            index = row["_original_index"]
            if index not in self.queue_manager.items:
                status = "completed" if row["_is_complete"] else "pending"
                self.queue_manager.add(QueueItem(original_index=index, status=status))
                changed = True
            elif row["_is_complete"] and self.queue_manager.items[index].status != "completed":
                item = self.queue_manager.items[index]
                item.status = "completed"
                item.leased_by_id = None
                item.leased_at = None
                changed = True
            elif not row["_is_complete"] and self.queue_manager.items[index].status == "completed":
                item = self.queue_manager.items[index]
                item.status = "pending"
                item.completed_by_id = None
                item.completed_at = None
                changed = True
        if changed:
            self.queue_manager.save()

    @property
    def is_complete(self) -> bool:
        items = self.queue_manager.get_all()
        return bool(items) and all(item.status == "completed" for item in items)

    @property
    def pending_count(self) -> int:
        return sum(
            item.status == "pending" for item in self.queue_manager.get_all()
        )


batch_contexts: Dict[str, BatchContext] = {}
batch_contexts_lock = threading.Lock()
_renaming_jobs: Dict[str, Dict[str, Any]] = {}
_renaming_jobs_lock = threading.Lock()
_renaming_clone_lock = threading.Lock()


def _batch_id(root: Path) -> str:
    return hashlib.sha256(str(root.resolve()).encode("utf-8")).hexdigest()[:16]


def discover_batches() -> Tuple[List[BatchContext], List[str]]:
    """Discover slide batches beneath immediate SS* directories."""
    base = Path(Config.LABEL_CHECK_BATCHES)
    warnings: List[str] = []
    discovered: List[BatchContext] = []
    try:
        scanner_dirs = sorted(
            (path for path in base.iterdir() if path.is_dir() and path.name.startswith("SS")),
            key=lambda path: path.name.lower(),
        )
    except OSError as exc:
        app.logger.warning("Label-check batch directory unavailable: %s", exc)
        return [], [f"Batch directory is unavailable: {base}"]

    candidates: List[Path] = []
    for scanner_dir in scanner_dirs:
        try:
            candidates.extend(
                sorted(
                    (path for path in scanner_dir.iterdir() if path.is_dir()),
                    key=lambda path: path.name.lower(),
                )
            )
        except OSError as exc:
            app.logger.warning("Scanner directory unavailable: %s", exc)
            warnings.append(f"Skipped {scanner_dir.name}: directory is unavailable.")

    for root in candidates:
        display_name = f"{root.parent.name}/{root.name}"
        missing = [
            name for name in ("label", "macro")
            if not (root / name).is_dir() or not os.access(root / name, os.R_OK | os.X_OK)
        ]
        csv_path = root / "enriched.csv"
        if not csv_path.is_file() or not os.access(csv_path, os.R_OK):
            missing.append("enriched.csv")
        if missing:
            warnings.append(f"Skipped {display_name}: missing {', '.join(missing)}.")
            continue
        try:
            batch_id = _batch_id(root)
            with batch_contexts_lock:
                context = batch_contexts.get(batch_id)
                if context is None or context.root.resolve() != root.resolve():
                    context = BatchContext(batch_id, root.resolve())
                    batch_contexts[batch_id] = context
            context.queue_manager.load()
            context.refresh()
            context.load_completed_stages(create_if_missing=True)
            if not context.data_manager.data:
                raise DataLoadError("enriched.csv has no slide rows")
            if "ParsingQCPassed" not in context.data_manager.headers:
                warnings.append(
                    f"Skipped {display_name}: enriched.csv is missing ParsingQCPassed."
                )
                continue
            discovered.append(context)
        except (DataLoadError, OSError, ValueError) as exc:
            app.logger.warning("Skipping invalid batch %s: %s", root, exc)
            warnings.append(f"Skipped {display_name}: {exc}")

    return discovered, warnings


def _selected_batch(allow_completed: bool = False) -> Tuple[Optional[BatchContext], List[BatchContext], List[str]]:
    batches, warnings = discover_batches()
    available = [batch for batch in batches if not batch.qc_complete]
    if request.args.get("choose") == "1":
        session.pop("qc_batch_id", None)
        return None, available, warnings
    requested_id = request.values.get("batch") or session.get("qc_batch_id")
    selected = next((batch for batch in batches if batch.id == requested_id), None)
    if selected and (allow_completed or not selected.qc_complete):
        session["qc_batch_id"] = selected.id
        return selected, available, warnings
    if requested_id:
        session.pop("qc_batch_id", None)
    return None, available, warnings


def _renaming_batches() -> Tuple[List[BatchContext], List[str]]:
    batches, warnings = discover_batches()
    return [
        batch for batch in batches
        if batch.completed_stages["QC"] and not batch.completed_stages["Renamed"]
    ], warnings


def _renaming_context(batch_id: str) -> Optional[BatchContext]:
    batches, _ = discover_batches()
    return next(
        (
            batch for batch in batches
            if batch.id == batch_id
            and batch.completed_stages["QC"]
            and not batch.completed_stages["Renamed"]
        ),
        None,
    )


def _renaming_job_state(batch_id: str) -> Dict[str, Any]:
    with _renaming_jobs_lock:
        return dict(_renaming_jobs.get(batch_id, {"status": "idle", "error": ""}))


def _start_renaming_job(
    context: BatchContext,
    *,
    old_accession: Optional[str] = None,
    new_accession: Optional[str] = None,
    force: bool = False,
) -> bool:
    """Start one background preparation or accession retry for a batch."""
    with _renaming_jobs_lock:
        existing = _renaming_jobs.get(context.id, {})
        if existing.get("status") in {"preparing", "retrying"}:
            return False
        if (
            not force
            and old_accession is None
            and (context.root / "name_mapping.csv").exists()
        ):
            _renaming_jobs[context.id] = {"status": "ready", "error": ""}
            return False
        _renaming_jobs[context.id] = {
            "status": "retrying" if old_accession else "preparing",
            "error": "",
        }

    def worker() -> None:
        try:
            with _renaming_clone_lock:
                if old_accession is not None and new_accession is not None:
                    renaming.retry_group(
                        context.root,
                        Path(Config.COPATH_CLONE),
                        Path(Config.LABEL_CHECK_BATCHES),
                        old_accession,
                        new_accession,
                    )
                else:
                    renaming.prepare_batch(
                        context.root,
                        Path(Config.COPATH_CLONE),
                        Path(Config.LABEL_CHECK_BATCHES),
                    )
            state = {"status": "ready", "error": ""}
        except Exception as exc:
            app.logger.exception("Renaming preparation failed for batch %s", context.id)
            state = {"status": "failed", "error": str(exc)}
        with _renaming_jobs_lock:
            _renaming_jobs[context.id] = state

    threading.Thread(target=worker, daemon=True).start()
    return True

# ==============================================================================
# 8. HELPER FUNCTIONS
# ==============================================================================
def _release_expired_leases(context: BatchContext):
    """Scans for and releases any item leases that have expired."""
    data_manager = context.data_manager
    queue_manager = context.queue_manager
    lease_duration = datetime.timedelta(seconds=app.config["LEASE_DURATION_SECONDS"])
    expired_time = datetime.datetime.utcnow() - lease_duration

    # Look for expired leases in QueueManager
    expired_items = [
        item for item in queue_manager.get_all()
        if item.status == "leased" and item.leased_at and item.leased_at < expired_time
    ]

    if expired_items:
        count = 0
        for item in expired_items:
            try:
                row = data_manager.get_row(item.original_index)
                acc_id = row.get("AccessionID", "Unknown") if row else "Unknown"
                
                app.logger.info(
                    f"Lease expired for item {item.original_index} ({acc_id}), leased by {item.leased_by_id}."
                )
                item.status = "pending"
                item.leased_by_id = None
                item.leased_at = None
                count += 1
            except Exception as e:
                app.logger.error(f"Error releasing lease for item {item.original_index}: {e}")
        
        if count > 0:
            queue_manager.save()
            flash(
                f"{count} item(s) had expired leases and were returned to the queue.",
                "warning",
            )


def _create_backup(context: BatchContext, suffix: str = "") -> None:
    """Creates a timestamped backup of the current CSV file."""
    source_path = str(context.csv_path)
    if not os.path.exists(source_path):
        return
    try:
        os.makedirs(Config.BACKUP_DIR, exist_ok=True)
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        name_part = f"{context.name}_{context.id}_enriched.csv_{timestamp}"
        if suffix:
            name_part += f"_{suffix}"
        backup_path = os.path.join(Config.BACKUP_DIR, f"{name_part}.bak")
        shutil.copy2(source_path, backup_path)
    except Exception as e:
        raise BackupError(f"Backup failed: {e}")


def _is_row_incomplete(row_dict: Dict[str, Any]) -> bool:
    return not row_dict.get("_is_complete", False)


def _qc_row_validation_errors(row: Dict[str, Any]) -> List[str]:
    """Return the QC fields that are missing or invalid for a completed row."""
    errors = []
    accession_id = str(row.get("AccessionID") or "").strip()
    if not accession_id:
        errors.append("Accession ID is required")
    elif not _accession_pattern.fullmatch(accession_id):
        errors.append("Accession ID must match A12-123")

    if not str(row.get("BlockNumber") or "").strip():
        errors.append("Block Number is required")
    if not str(row.get("Stain") or "").strip():
        errors.append("Stain is required")
    return errors


def _requeue_invalid_qc_rows(context: BatchContext) -> List[int]:
    """Return invalid completed rows to the pending queue."""
    invalid_indices = []
    for row in context.data_manager.data:
        if not _qc_row_validation_errors(row):
            continue

        index = row["_original_index"]
        invalid_indices.append(index)
        row["_is_complete"] = False

        item = context.queue_manager.get(index)
        if item is None:
            item = QueueItem(original_index=index)
            context.queue_manager.add(item)
        item.status = "pending"
        item.leased_by_id = None
        item.leased_at = None
        item.completed_by_id = None
        item.completed_at = None

    if invalid_indices:
        context.queue_manager.save()
    return invalid_indices


def flash_messages() -> List[Dict[str, str]]:
    return [
        {"category": category, "message": message}
        for category, message in get_flashed_messages(with_categories=True)
    ]


# ==============================================================================
# SLIDE DIGITIZATION LOG HELPERS
# ==============================================================================
SDL_HEADERS = (
    "Accession ID",
    "Organ",
    "Type",
    "Slides Count",
    "Scanner",
    "Carousel Rack",
    "Date Loaded",
    "Time Loaded",
    "Date Unloaded",
    "Time Unloaded",
    "Ran Label-Check",
    "Finished QC",
    "Collected CoPath Data",
    "Renamed",
    "Pushed to SFTP Server",
    "Notes",
)

SDL_STATUS_HEADERS = (
    "Ran Label-Check",
    "Finished QC",
    "Collected CoPath Data",
    "Renamed",
    "Pushed to SFTP Server",
)

SDL_FORM_FIELDS = {
    "Accession ID": "accession_id",
    "Organ": "organ",
    "Type": "type",
    "Slides Count": "slides_count",
    "Scanner": "scanner",
    "Carousel Rack": "carousel_rack",
    "Date Loaded": "date_loaded",
    "Time Loaded": "time_loaded",
    "Date Unloaded": "date_unloaded",
    "Time Unloaded": "time_unloaded",
    "Notes": "notes",
}

_sdl_workbook_lock = threading.Lock()
_accession_pattern = re.compile(r"^[A-Z]{1,3}[0-9]{2}-[0-9]+$")
_rack_pattern = re.compile(r"^[0-9]+(?:\s*,\s*[0-9]+)*$")
_time_pattern = re.compile(r"^[0-9]{2}:[0-9]{2}$")


def _save_sdl_workbook(workbook) -> None:
    """Atomically replaces the SDL workbook with the supplied workbook."""
    workbook_path = Config.SDL_FILE_PATH
    workbook_dir = os.path.dirname(workbook_path)
    file_mode = stat.S_IMODE(os.stat(workbook_path).st_mode)
    file_descriptor, temporary_path = tempfile.mkstemp(
        prefix=".Slide_Digitization_Log.", suffix=".xlsx", dir=workbook_dir
    )
    os.close(file_descriptor)
    try:
        workbook.save(temporary_path)
        os.chmod(temporary_path, file_mode)
        os.replace(temporary_path, workbook_path)
    finally:
        if os.path.exists(temporary_path):
            os.remove(temporary_path)


def _load_sdl_workbook():
    """Loads and validates the configured SDL workbook and worksheet."""
    workbook_path = Config.SDL_FILE_PATH
    if not os.path.isfile(workbook_path):
        raise SDLWorkbookError(
            f"Slide Digitization Log not found at {workbook_path}."
        )

    try:
        workbook = load_workbook(workbook_path)
    except Exception as exc:
        raise SDLWorkbookError(f"The Slide Digitization Log could not be opened: {exc}") from exc

    if Config.SDL_SHEET_NAME not in workbook.sheetnames:
        workbook.close()
        raise SDLWorkbookError(
            f"The workbook must contain a worksheet named '{Config.SDL_SHEET_NAME}'."
        )

    worksheet = workbook[Config.SDL_SHEET_NAME]
    has_any_value = any(
        cell.value is not None
        for row in worksheet.iter_rows()
        for cell in row
    )
    initialized_headers = False
    if not has_any_value:
        worksheet.append(SDL_HEADERS)
        initialized_headers = True
    else:
        actual_headers = tuple(
            str(worksheet.cell(row=1, column=column).value).strip()
            if worksheet.cell(row=1, column=column).value is not None
            else ""
            for column in range(1, len(SDL_HEADERS) + 1)
        )
        extra_headers = [
            worksheet.cell(row=1, column=column).value
            for column in range(len(SDL_HEADERS) + 1, worksheet.max_column + 1)
            if worksheet.cell(row=1, column=column).value is not None
        ]
        if actual_headers != SDL_HEADERS or extra_headers:
            workbook.close()
            raise SDLWorkbookError(
                "The SDL worksheet headers do not match the required schema."
            )

    return workbook, worksheet, initialized_headers


def _coerce_sdl_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value != 0
    if isinstance(value, str):
        return value.strip().lower() in {"true", "yes", "y", "1"}
    return False


def _format_sdl_value(header: str, value: Any) -> str:
    if value is None:
        return ""
    if header.startswith("Date ") and isinstance(value, (datetime.datetime, datetime.date)):
        return value.strftime("%Y-%m-%d")
    if header.startswith("Time ") and isinstance(value, (datetime.datetime, datetime.time)):
        return value.strftime("%H:%M")
    return str(value)


def _sdl_row_signature(worksheet: Worksheet, row_number: int) -> str:
    values = tuple(
        worksheet.cell(row=row_number, column=column).value
        for column in range(1, len(SDL_HEADERS) + 1)
    )
    return hashlib.sha256(repr(values).encode("utf-8")).hexdigest()


def _read_sdl_rows(worksheet: Worksheet) -> List[Dict[str, Any]]:
    rows = []
    for row_number in range(2, worksheet.max_row + 1):
        raw_values = {
            header: worksheet.cell(row=row_number, column=column).value
            for column, header in enumerate(SDL_HEADERS, start=1)
        }
        if all(value is None for value in raw_values.values()):
            continue
        rows.append(
            {
                "worksheet_row": row_number,
                "values": {
                    header: _format_sdl_value(header, value)
                    for header, value in raw_values.items()
                },
                "statuses": {
                    header: _coerce_sdl_bool(raw_values[header])
                    for header in SDL_STATUS_HEADERS
                },
                "signature": _sdl_row_signature(worksheet, row_number),
            }
        )
    return rows


def _submitted_sdl_form() -> Dict[str, str]:
    return {
        header: request.form.get(field_name, "").strip()
        for header, field_name in SDL_FORM_FIELDS.items()
    }


def _validate_sdl_form(values: Dict[str, str]) -> Dict[str, Any]:
    accession_id = values["Accession ID"]
    if not _accession_pattern.fullmatch(accession_id):
        raise SDLValidationError(
            "Accession ID must match the format A12-123 (1-3 uppercase letters, "
            "2 digits, a hyphen, and one or more digits)."
        )
    if values["Organ"] not in Config.SDL_ORGANS:
        raise SDLValidationError("Select a valid Organ.")
    if not values["Type"]:
        raise SDLValidationError("Type is required.")
    try:
        slides_count = int(values["Slides Count"])
    except ValueError as exc:
        raise SDLValidationError("Slides Count must be an integer.") from exc
    if slides_count < 1:
        raise SDLValidationError("Slides Count must be at least 1.")
    if values["Scanner"] not in Config.SDL_SCANNERS:
        raise SDLValidationError("Select a valid Scanner.")

    carousel_rack = values["Carousel Rack"]
    if not _rack_pattern.fullmatch(carousel_rack):
        raise SDLValidationError(
            "Carousel Rack must contain positive integers separated by commas."
        )
    rack_numbers = [int(value.strip()) for value in carousel_rack.split(",")]
    if any(value < 1 for value in rack_numbers):
        raise SDLValidationError("Carousel Rack numbers must be at least 1.")

    try:
        date_loaded = datetime.date.fromisoformat(values["Date Loaded"])
    except ValueError as exc:
        raise SDLValidationError("Date Loaded must use YYYY-MM-DD format.") from exc
    if not _time_pattern.fullmatch(values["Time Loaded"]):
        raise SDLValidationError("Time Loaded must use HH:MM 24-hour format.")
    try:
        time_loaded = datetime.time.fromisoformat(values["Time Loaded"])
    except ValueError as exc:
        raise SDLValidationError("Time Loaded must be a valid 24-hour time.") from exc

    date_unloaded_value = values["Date Unloaded"]
    time_unloaded_value = values["Time Unloaded"]
    if bool(date_unloaded_value) != bool(time_unloaded_value):
        raise SDLValidationError(
            "Date Unloaded and Time Unloaded must either both be supplied or both be blank."
        )
    date_unloaded = None
    time_unloaded = None
    if date_unloaded_value:
        try:
            date_unloaded = datetime.date.fromisoformat(date_unloaded_value)
        except ValueError as exc:
            raise SDLValidationError("Date Unloaded must use YYYY-MM-DD format.") from exc
        if not _time_pattern.fullmatch(time_unloaded_value):
            raise SDLValidationError("Time Unloaded must use HH:MM 24-hour format.")
        try:
            time_unloaded = datetime.time.fromisoformat(time_unloaded_value)
        except ValueError as exc:
            raise SDLValidationError("Time Unloaded must be a valid 24-hour time.") from exc
        if datetime.datetime.combine(date_unloaded, time_unloaded) < datetime.datetime.combine(
            date_loaded, time_loaded
        ):
            raise SDLValidationError("The unloaded timestamp cannot precede the loaded timestamp.")

    return {
        "Accession ID": accession_id,
        "Organ": values["Organ"],
        "Type": values["Type"],
        "Slides Count": slides_count,
        "Scanner": values["Scanner"],
        "Carousel Rack": ", ".join(str(value) for value in rack_numbers),
        "Date Loaded": date_loaded,
        "Time Loaded": time_loaded,
        "Date Unloaded": date_unloaded,
        "Time Unloaded": time_unloaded,
        "Notes": values["Notes"],
    }


def _render_sdl_page(
    form_values: Optional[Dict[str, str]] = None,
    edit_row: Optional[int] = None,
    edit_signature: str = "",
):
    workbook = None
    try:
        with _sdl_workbook_lock:
            workbook, worksheet, initialized_headers = _load_sdl_workbook()
            if initialized_headers:
                _save_sdl_workbook(workbook)
            rows = _read_sdl_rows(worksheet)
    except SDLWorkbookError as exc:
        return render_template(
            "sdl.html",
            workbook_available=False,
            workbook_error=str(exc),
            messages=flash_messages(),
        )
    except Exception as exc:
        app.logger.exception("Unexpected error while reading the SDL workbook")
        return render_template(
            "sdl.html",
            workbook_available=False,
            workbook_error=f"The Slide Digitization Log could not be read: {exc}",
            messages=flash_messages(),
        )
    finally:
        if workbook is not None:
            workbook.close()

    if edit_row is not None and form_values is None:
        selected_row = next(
            (row for row in rows if row["worksheet_row"] == edit_row), None
        )
        if selected_row is None:
            flash("The selected SDL row no longer exists.", "error")
            return redirect(url_for("sdl"))
        form_values = {
            header: selected_row["values"][header]
            for header in SDL_FORM_FIELDS
        }
        edit_signature = selected_row["signature"]

    return render_template(
        "sdl.html",
        workbook_available=True,
        rows=rows,
        headers=SDL_HEADERS,
        status_headers=SDL_STATUS_HEADERS,
        form_fields=SDL_FORM_FIELDS,
        form_values=form_values or {header: "" for header in SDL_FORM_FIELDS},
        edit_row=edit_row,
        edit_signature=edit_signature,
        organ_options=Config.SDL_ORGANS,
        scanner_options=Config.SDL_SCANNERS,
        messages=flash_messages(),
    )


def _read_inventory_page(
    inventory_path: Path, requested_page: int, rows_per_page: int = 100
) -> Tuple[List[str], List[List[str]], int, int, int]:
    """Read one page of a headered inventory CSV without retaining the whole file."""
    headers: List[str] = []
    page_rows: List[List[str]] = []
    total_rows = 0

    try:
        with inventory_path.open("r", encoding="utf-8-sig", newline="") as inventory_file:
            reader = csv.reader(inventory_file, strict=True)
            try:
                headers = next(reader)
            except StopIteration:
                raise InventoryReadError("This inventory is empty and has no header row.")
            if not headers:
                raise InventoryReadError("This inventory does not contain a usable header row.")

            page_start = (requested_page - 1) * rows_per_page
            page_end = page_start + rows_per_page
            for row_number, row in enumerate(reader):
                if page_start <= row_number < page_end:
                    page_rows.append(row)
                total_rows += 1
    except InventoryReadError:
        raise
    except UnicodeDecodeError as exc:
        raise InventoryReadError("This inventory is not valid UTF-8 text.") from exc
    except csv.Error as exc:
        raise InventoryReadError(f"This inventory contains invalid CSV data: {exc}") from exc
    except OSError as exc:
        raise InventoryReadError(f"This inventory could not be read: {exc}") from exc

    total_pages = max(1, (total_rows + rows_per_page - 1) // rows_per_page)
    current_page = min(requested_page, total_pages)

    if current_page != requested_page:
        return _read_inventory_page(inventory_path, current_page, rows_per_page)

    return headers, page_rows, total_rows, current_page, total_pages


# ==============================================================================
# 9. PIPELINE LAUNCHER
# ==============================================================================
PIPELINE_FORM_DEFAULTS = {
    "input_dir": "",
    "output_dir": "",
    "start_from": "1",
    "end_at": "3",
    "input_mode": "auto",
    "macro_workers": "4",
    "macro_extensions": "svs",
    "macro_image_extensions": "png, jpg, jpeg, tif, tiff, bmp",
    "thumbnail_width": "300",
    "thumbnail_height": "300",
    "ocr_workers": "4",
    "ocr_use_cpu": "",
    "naming_accession_pattern": r"\b([A-Za-z]{1,3}\s*\d{2}\s*[ -/]\s*\d+)\b",
    "naming_workers": "4",
}


class PipelineJob:
    """In-memory state for one pipeline child process."""

    def __init__(
        self,
        job_id: str,
        owner_id: str,
        process: subprocess.Popen,
        output_path: Optional[str] = None,
    ):
        self.id = job_id
        self.owner_id = owner_id
        self.process = process
        self.status = "running"
        self.return_code: Optional[int] = None
        self.output = ""
        self.output_path = output_path


_pipeline_jobs: Dict[str, PipelineJob] = {}
_pipeline_jobs_lock = threading.Lock()
_pipeline_active_job_id: Optional[str] = None


def _pipeline_form_values(source=None) -> Dict[str, str]:
    values = dict(PIPELINE_FORM_DEFAULTS)
    if source is not None:
        for key in values:
            submitted = source.get(key)
            if submitted is not None:
                values[key] = submitted
    return values


def _positive_pipeline_integer(
    values: Dict[str, str], field: str, label: str, errors: List[str]
) -> Optional[int]:
    try:
        value = int(values[field])
        if value <= 0:
            raise ValueError
        return value
    except (TypeError, ValueError):
        errors.append(f"{label} must be a positive whole number.")
        return None


def _pipeline_extensions(value: str, label: str, errors: List[str]) -> List[str]:
    extensions = [item.lstrip(".") for item in re.split(r"[\s,]+", value.strip()) if item]
    if not extensions:
        errors.append(f"{label} must contain at least one extension.")
    elif any(not re.fullmatch(r"[A-Za-z0-9.]+", item) for item in extensions):
        errors.append(f"{label} may contain only letters, numbers, and periods.")
    return extensions


def _pipeline_command(values: Dict[str, str]) -> Tuple[Optional[List[str]], List[str]]:
    errors: List[str] = []
    input_text = values["input_dir"].strip()
    output_text = values["output_dir"].strip()
    if not input_text:
        errors.append("Input directory is required.")
    if not output_text:
        errors.append("Output directory is required.")

    input_dir = Path(input_text).expanduser().resolve() if input_text else None
    output_dir = Path(output_text).expanduser().resolve() if output_text else None
    if input_dir is not None and not input_dir.is_dir():
        errors.append("Input directory must be an existing directory on the server.")
    if output_dir is not None and output_dir.exists() and not output_dir.is_dir():
        errors.append("Output directory points to a file, not a directory.")

    try:
        start_stage = int(values["start_from"])
        end_stage = int(values["end_at"])
        if start_stage not in (1, 2, 3) or end_stage not in (1, 2, 3):
            raise ValueError
        if end_stage < start_stage:
            errors.append("End at cannot be earlier than Start from.")
    except (TypeError, ValueError):
        start_stage = end_stage = 0
        errors.append("Choose valid starting and ending stages.")

    macro_workers = _positive_pipeline_integer(
        values, "macro_workers", "Get macro workers", errors
    )
    thumbnail_width = _positive_pipeline_integer(
        values, "thumbnail_width", "Thumbnail width", errors
    )
    thumbnail_height = _positive_pipeline_integer(
        values, "thumbnail_height", "Thumbnail height", errors
    )
    ocr_workers = _positive_pipeline_integer(
        values, "ocr_workers", "Dual OCR workers", errors
    )
    naming_workers = _positive_pipeline_integer(
        values, "naming_workers", "Name files workers", errors
    )
    macro_extensions = _pipeline_extensions(
        values["macro_extensions"], "Slide extensions", errors
    )
    image_extensions = _pipeline_extensions(
        values["macro_image_extensions"], "Image extensions", errors
    )

    if values["input_mode"] not in ("auto", "slides", "images"):
        errors.append("Choose a valid input mode.")
    try:
        re.compile(values["naming_accession_pattern"])
    except re.error as exc:
        errors.append(f"Accession pattern is not a valid regular expression: {exc}")

    if output_dir is not None and start_stage == 2:
        if not (output_dir / "slide_mapping.csv").is_file():
            errors.append(
                "Starting from Dual OCR requires slide_mapping.csv in the output directory."
            )
    if output_dir is not None and start_stage == 3:
        if not (output_dir / "ocr.csv").is_file():
            errors.append(
                "Starting from Name files requires ocr.csv in the output directory."
            )

    pipeline_script = Path(__file__).resolve().with_name("pipeline.py")
    if not pipeline_script.is_file():
        errors.append(f"Pipeline script is unavailable: {pipeline_script}")
    if errors:
        return None, errors

    command = [
        sys.executable,
        "-u",
        str(pipeline_script),
        "--input-dir",
        str(input_dir),
        "--output-dir",
        str(output_dir),
        "--start-from",
        str(start_stage),
        "--end-at",
        str(end_stage),
        "--input-mode",
        values["input_mode"],
        "--macro-workers",
        str(macro_workers),
        "--macro-extensions",
        *macro_extensions,
        "--macro-image-extensions",
        *image_extensions,
        "--macro-thumbnail-size",
        str(thumbnail_width),
        str(thumbnail_height),
        "--ocr-workers",
        str(ocr_workers),
        "--naming-accession-pattern",
        values["naming_accession_pattern"],
        "--naming-workers",
        str(naming_workers),
    ]
    if values["ocr_use_cpu"] == "on":
        command.append("--ocr-use-cpu")
    return command, []


def _read_pipeline_output(job: PipelineJob) -> None:
    """Drain merged child output and finalize job state."""
    global _pipeline_active_job_id
    decoder = codecs.getincrementaldecoder("utf-8")(errors="replace")
    try:
        if job.process.stdout is not None:
            while True:
                chunk = job.process.stdout.read(4096)
                if not chunk:
                    break
                text = decoder.decode(chunk)
                if text:
                    with _pipeline_jobs_lock:
                        job.output += text
                    if job.output_path:
                        with open(job.output_path, "a", encoding="utf-8") as output_file:
                            output_file.write(text)
            trailing_text = decoder.decode(b"", final=True)
            if trailing_text:
                with _pipeline_jobs_lock:
                    job.output += trailing_text
                if job.output_path:
                    with open(job.output_path, "a", encoding="utf-8") as output_file:
                        output_file.write(trailing_text)
        return_code = job.process.wait()
        with _pipeline_jobs_lock:
            job.return_code = return_code
            job.status = "succeeded" if return_code == 0 else "failed"
        if job.output_path:
            api_store.update_job(
                job.id,
                status=job.status,
                return_code=return_code,
                completed_at=_iso_utc(),
            )
    except Exception as exc:
        app.logger.exception("Failed while reading pipeline output")
        with _pipeline_jobs_lock:
            job.output += f"\nLauncher error while reading output: {exc}\n"
            job.return_code = job.process.poll()
            job.status = "failed"
        if job.output_path:
            with open(job.output_path, "a", encoding="utf-8") as output_file:
                output_file.write(f"\nLauncher error while reading output: {exc}\n")
            api_store.update_job(
                job.id,
                status="failed",
                return_code=job.return_code,
                completed_at=_iso_utc(),
            )
    finally:
        with _pipeline_jobs_lock:
            if _pipeline_active_job_id == job.id:
                _pipeline_active_job_id = None


def _start_pipeline_job(
    command: List[str],
    owner_id: str,
    *,
    job_id: Optional[str] = None,
    request_values: Optional[Dict[str, Any]] = None,
    token_id: Optional[str] = None,
    idempotency_key: Optional[str] = None,
    payload_hash: Optional[str] = None,
) -> PipelineJob:
    global _pipeline_active_job_id
    with _pipeline_jobs_lock:
        if _pipeline_active_job_id is not None:
            active_job = _pipeline_jobs.get(_pipeline_active_job_id)
            if active_job is not None and active_job.status == "running":
                raise RuntimeError("Another Label-Check pipeline is already running.")
            _pipeline_active_job_id = None

        resolved_job_id = job_id or str(uuid.uuid4())
        try:
            output_path = api_store.reserve_job(
                resolved_job_id,
                owner_id,
                request_values or {},
                command,
                token_id,
                idempotency_key,
                payload_hash,
            )
        except sqlite3.IntegrityError as exc:
            raise RuntimeError("Another Label-Check pipeline is already running.") from exc

        try:
            process = subprocess.Popen(
                command,
                cwd=Path(__file__).resolve().parent.parent,
                stdin=subprocess.DEVNULL,
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                bufsize=0,
                env={**os.environ, "PYTHONUNBUFFERED": "1"},
            )
        except OSError:
            api_store.update_job(
                resolved_job_id, status="failed", completed_at=_iso_utc()
            )
            raise
        job = PipelineJob(resolved_job_id, owner_id, process, output_path)
        _pipeline_jobs[job.id] = job
        _pipeline_active_job_id = job.id
        api_store.update_job(
            job.id,
            status="running",
            started_at=_iso_utc(),
            launcher_pid=os.getpid(),
        )

    reader = threading.Thread(target=_read_pipeline_output, args=(job,), daemon=True)
    reader.start()
    return job


def _pipeline_job_for_user(job_id: str) -> Optional[PipelineJob]:
    with _pipeline_jobs_lock:
        job = _pipeline_jobs.get(job_id)
        if job is None:
            return None
        if job.owner_id != str(current_user.id) and not current_user.is_admin:
            return None
        return job


def _pipeline_is_busy() -> bool:
    with _pipeline_jobs_lock:
        if _pipeline_active_job_id is None:
            return False
        job = _pipeline_jobs.get(_pipeline_active_job_id)
        return job is not None and job.status == "running"


def csrf_token() -> str:
    token = session.get("_csrf_token")
    if not token:
        token = secrets.token_urlsafe(32)
        session["_csrf_token"] = token
    return token


app.jinja_env.globals["csrf_token"] = csrf_token


def _api_problem(status: int, code: str, title: str, detail: str):
    response = jsonify(
        {
            "type": f"https://label-check.invalid/problems/{code}",
            "status": status,
            "code": code,
            "title": title,
            "detail": detail,
            "request_id": getattr(g, "request_id", uuid.uuid4().hex),
        }
    )
    response.status_code = status
    response.content_type = "application/problem+json"
    if status == 401:
        response.headers["WWW-Authenticate"] = 'Bearer realm="label-check-api"'
    return response


def _require_api_scope(scope: str, bucket: str = "read"):
    def decorator(view):
        @functools.wraps(view)
        def wrapped(*args, **kwargs):
            if app.config.get("API_REQUIRE_HTTPS", True) and not app.testing and not request.is_secure:
                return _api_problem(400, "https_required", "HTTPS required", "The API is available only over HTTPS.")
            header = request.headers.get("Authorization", "")
            if not header.startswith("Bearer ") or header.count(" ") != 1:
                return _api_problem(401, "invalid_token", "Authentication required", "Provide a valid bearer token.")
            token = api_store.authenticate_token(header[7:])
            user = user_manager.get(token["user_id"]) if token else None
            if token is None or user is None:
                return _api_problem(401, "invalid_token", "Authentication required", "The bearer token is invalid, expired, or revoked.")
            g.api_token = token
            g.api_user = user
            if scope not in token["scopes"]:
                return _api_problem(403, "insufficient_scope", "Insufficient scope", f"This endpoint requires the {scope} scope.")
            limit = (
                app.config["API_SUBMIT_RATE_LIMIT"]
                if bucket == "submit"
                else app.config["API_READ_RATE_LIMIT"]
            )
            allowed, remaining, retry_after = api_store.rate_limit(
                token["token_id"], bucket, limit, app.config["API_RATE_WINDOW_SECONDS"]
            )
            g.rate_limit = limit
            g.rate_remaining = remaining
            g.rate_retry_after = retry_after
            if not allowed:
                response = _api_problem(429, "rate_limit_exceeded", "Rate limit exceeded", "Retry after the current rate-limit window.")
                response.headers["Retry-After"] = str(retry_after)
                return response
            return view(*args, **kwargs)
        return wrapped
    return decorator


def _api_job_document(record: Dict[str, Any]) -> Dict[str, Any]:
    job_id = record["job_id"]
    return {
        "id": job_id,
        "status": record["status"],
        "created_at": record["created_at"],
        "started_at": record.get("started_at"),
        "completed_at": record.get("completed_at"),
        "return_code": record.get("return_code"),
        "links": {
            "self": url_for("api_pipeline_job", job_id=job_id, _external=True),
            "output": url_for("api_pipeline_job_output", job_id=job_id, _external=True),
        },
    }


@app.after_request
def api_response_metadata(response):
    if request.path.startswith("/api/v1/"):
        request_id = getattr(g, "request_id", uuid.uuid4().hex)
        response.headers["X-Request-ID"] = request_id
        if hasattr(g, "rate_limit"):
            response.headers["X-RateLimit-Limit"] = str(g.rate_limit)
            response.headers["X-RateLimit-Remaining"] = str(g.rate_remaining)
        token = getattr(g, "api_token", None)
        app.logger.info(
            "API_AUDIT %s",
            json.dumps(
                {
                    "request_id": request_id,
                    "token_id": token.get("token_id") if token else None,
                    "user_id": token.get("user_id") if token else None,
                    "method": request.method,
                    "path": request.path,
                    "status": response.status_code,
                    "remote_addr": request.remote_addr,
                    "idempotency_key": request.headers.get("Idempotency-Key"),
                },
                sort_keys=True,
            ),
        )
    return response


@app.errorhandler(404)
def api_not_found(error):
    if request.path.startswith("/api/v1/"):
        return _api_problem(404, "not_found", "Not found", "The requested API resource does not exist.")
    return error


@app.errorhandler(405)
def api_method_not_allowed(error):
    if request.path.startswith("/api/v1/"):
        return _api_problem(405, "method_not_allowed", "Method not allowed", "This API resource does not support the requested method.")
    return error


@app.errorhandler(500)
def api_internal_error(error):
    if request.path.startswith("/api/v1/"):
        return _api_problem(500, "internal_error", "Internal server error", "The request could not be completed.")
    return error


# ==============================================================================
# 10. FLASK ROUTES
# ==============================================================================
@app.before_request
def before_request_handler():
    if request.path.startswith("/api/v1/"):
        supplied_request_id = request.headers.get("X-Request-ID", "")
        g.request_id = (
            supplied_request_id
            if re.fullmatch(r"[A-Za-z0-9._-]{1,64}", supplied_request_id)
            else uuid.uuid4().hex
        )
        return None

    if (
        request.method in {"POST", "PUT", "PATCH", "DELETE"}
        and app.config.get("CSRF_ENABLED", True)
        and not app.testing
    ):
        expected = session.get("_csrf_token", "")
        supplied = request.form.get("csrf_token", "") or request.headers.get("X-CSRF-Token", "")
        if not expected or not supplied or not hmac.compare_digest(expected, supplied):
            return "Invalid or missing CSRF token.", 400

    if request.endpoint in [
        "static",
        "serve_relative_image",
        "login",
        "logout",
        "sdl",
        "inventories",
    ]:
        return

    session.setdefault("show_only_incomplete", False)


@app.route("/login", methods=["GET", "POST"])
def login():
    if current_user.is_authenticated:
        return redirect(url_for("index"))
    
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        # Use UserManager
        user = user_manager.get(username)
        
        if user and user.verify_password(password):
            login_user(user)
            app.logger.info(f"User '{username}' logged in successfully.")
            return redirect(request.args.get("next") or url_for("index"))
        
        flash("Invalid username or password.", "error")
        
    return render_template("login.html", messages=flash_messages())


@app.route("/logout")
@login_required
def logout():
    logout_user()
    flash("You have been logged out.", "info")
    return redirect(url_for("login"))


@app.route("/users")
@login_required
def users_management():
    if not current_user.is_admin:
        flash("You do not have permission to access this page.", "error")
        return redirect(url_for("index"))
        
    users = user_manager.get_all()
    return render_template("users.html", users=users, messages=flash_messages())


@app.route("/add_user", methods=["POST"])
@login_required
def add_user():
    if not current_user.is_admin:
        return redirect(url_for("index"))
        
    username = request.form.get("username", "").strip()
    password = request.form.get("password", "").strip()
    
    if not username or not password:
        flash("Username and password are required.", "error")
        return redirect(url_for("users_management"))
        
    if user_manager.get(username):
        flash(f"User '{username}' already exists.", "error")
        return redirect(url_for("users_management"))
        
    try:
        u = User(id=username, password_hash="", is_admin=(request.form.get("is_admin") == "on"))
        u.set_password(password)
        user_manager.add(u)
        flash(f"User '{username}' created successfully.", "success")
    except Exception as e:
        flash(f"An error occurred while adding the user: {e}", "error")
        
    return redirect(url_for("users_management"))


@app.route("/", methods=["GET"])
@login_required
def index():
    requested_index = request.args.get("index")
    if requested_index is not None:
        return redirect(url_for("qc", index=requested_index))

    return render_template("index.html", messages=flash_messages())


@app.route("/renaming", methods=["GET"])
@login_required
def renaming_page():
    batches, discovery_warnings = _renaming_batches()
    if request.args.get("choose") == "1":
        session.pop("renaming_batch_id", None)
    requested = request.args.get("batch") or session.get("renaming_batch_id")
    context = next((batch for batch in batches if batch.id == requested), None)
    if context is None:
        if requested:
            session.pop("renaming_batch_id", None)
        for batch in batches:
            if not (batch.root / "name_mapping.csv").exists():
                _start_renaming_job(batch)
        return render_template(
            "renaming.html",
            batches=batches,
            context=None,
            discovery_warnings=discovery_warnings,
            messages=flash_messages(),
            job_states={batch.id: _renaming_job_state(batch.id) for batch in batches},
        )

    session["renaming_batch_id"] = context.id
    mapping_path = context.root / "name_mapping.csv"
    if not mapping_path.exists():
        _start_renaming_job(context)
        return render_template(
            "renaming.html", batches=batches, context=context, groups=[], signature="",
            discovery_warnings=discovery_warnings, messages=flash_messages(),
            job_state=_renaming_job_state(context.id),
        )
    try:
        _, rows = renaming.read_csv(mapping_path)
        reports = renaming.report_rows(context.root, Path(Config.COPATH_CLONE))
        groups = renaming.group_mapping(rows, reports)
        signature = renaming.mapping_signature(rows)
    except renaming.RenamingError as exc:
        flash(str(exc), "error")
        groups, signature = [], ""
    return render_template(
        "renaming.html", batches=batches, context=context, groups=groups,
        signature=signature, discovery_warnings=discovery_warnings,
        messages=flash_messages(), job_state=_renaming_job_state(context.id),
    )


@app.route("/renaming/status/<batch_id>", methods=["GET"])
@login_required
def renaming_status(batch_id: str):
    context = _renaming_context(batch_id)
    if context is None:
        return jsonify({"status": "unavailable", "error": "Batch not found."}), 404
    state = _renaming_job_state(batch_id)
    state["ready"] = (context.root / "name_mapping.csv").exists()
    return jsonify(state)


@app.route("/renaming/prepare/<batch_id>", methods=["POST"])
@login_required
def renaming_prepare(batch_id: str):
    context = _renaming_context(batch_id)
    if context is None:
        flash("The batch is no longer available for renaming.", "warning")
        return redirect(url_for("renaming_page"))
    _start_renaming_job(context, force=True)
    flash("CoPath preparation started.", "info")
    return redirect(url_for("renaming_page", batch=batch_id))


@app.route("/renaming/retry/<batch_id>", methods=["POST"])
@login_required
def renaming_retry(batch_id: str):
    context = _renaming_context(batch_id)
    if context is None:
        flash("The batch is no longer available for renaming.", "warning")
        return redirect(url_for("renaming_page"))
    old_accession = request.form.get("old_accession", "").strip().upper()
    new_accession = request.form.get("accession_id", "").strip().upper()
    if not renaming.ACCESSION_RE.fullmatch(new_accession):
        flash("AccessionID must match A12-123.", "error")
    elif _start_renaming_job(
        context, old_accession=old_accession, new_accession=new_accession, force=True
    ):
        flash(f"CoPath retry started for {new_accession}.", "info")
    else:
        flash("A CoPath job is already running for this batch.", "warning")
    return redirect(url_for("renaming_page", batch=batch_id))


@app.route("/renaming/approve/<batch_id>", methods=["POST"])
@login_required
def renaming_approve(batch_id: str):
    context = _renaming_context(batch_id)
    if context is None:
        flash("The batch is no longer available for renaming.", "warning")
        return redirect(url_for("renaming_page"))
    mapping_path = context.root / "name_mapping.csv"
    old_accession = request.form.get("old_accession", "").strip().upper()
    values = {
        "AccessionID": request.form.get("accession_id", "").strip().upper(),
        "Organ": request.form.get("organ", "").strip().upper(),
        "PID": request.form.get("pid", "").strip().upper(),
        "AccessionDate": request.form.get("accession_date", "").strip().upper(),
        "Timepoint": request.form.get("timepoint", "").strip().upper(),
        "ImageType": request.form.get("image_type", "").strip().upper(),
        "SampAcqType": request.form.get("samp_acq_type", "").strip().upper(),
    }
    try:
        if _renaming_job_state(batch_id).get("status") in {"preparing", "retrying"}:
            raise renaming.RenamingError(
                "Wait for the active CoPath job to finish before approving names"
            )
        _, current_rows = renaming.read_csv(mapping_path)
        target_exists = any(
            row["AccessionID"] == values["AccessionID"]
            for row in current_rows
            if row["AccessionID"] != old_accession
        )
        if values["AccessionID"] != old_accession and not target_exists:
            raise renaming.RenamingError(
                "Retry CoPath after changing an accession ID before approving it"
            )
        if not target_exists:
            reports = renaming.report_rows(context.root, Path(Config.COPATH_CLONE))
            mrn = reports.get(old_accession, {}).get("mrn", "").strip()
            pid_error = renaming.validate_pid_assignment(
                Path(Config.COPATH_CLONE), values["Organ"], values["PID"], mrn
            )
            if pid_error:
                raise renaming.RenamingError(pid_error)
        slide_values: Dict[str, Dict[str, str]] = {}
        try:
            slide_count = int(request.form.get("slide_count", "0"))
        except ValueError as exc:
            raise renaming.RenamingError("Invalid slide submission") from exc
        for index in range(slide_count):
            path = request.form.get(f"original_path_{index}", "")
            slide_values[path] = {
                "Stain": request.form.get(f"stain_{index}", "").strip(),
                "BlockNumber": request.form.get(f"block_number_{index}", "").strip(),
                "SectionCount": request.form.get(f"section_count_{index}", "").strip(),
            }
        updated, merged = renaming.update_group(
            mapping_path, old_accession, values, slide_values,
            request.form.get("mapping_signature", ""),
        )
        if merged:
            flash("Accessions were merged. Review and approve the combined group.", "info")
        else:
            flash(f"Approved names for {values['AccessionID']}.", "success")
        if updated and all(renaming.parse_bool(row["Approved"]) for row in updated):
            with _renaming_clone_lock:
                renaming.finalize_batch(context.root, Path(Config.COPATH_CLONE))
                context.mark_renamed_complete()
            flash("All names are approved and the batch has been finalized.", "success")
            return redirect(url_for("renaming_page"))
    except (renaming.RenamingError, DataSaveError) as exc:
        app.logger.warning("Renaming approval failed for %s: %s", batch_id, exc)
        flash(str(exc), "error")
    except Exception as exc:
        app.logger.exception("Renaming finalization failed for %s", batch_id)
        flash(f"Renaming finalization failed: {exc}", "error")
    return redirect(url_for("renaming_page", batch=batch_id))


@app.route("/pipeline", methods=["GET"])
@login_required
def pipeline_launcher():
    job = None
    job_id = session.get("pipeline_job_id")
    if job_id:
        job = _pipeline_job_for_user(job_id)
    return render_template(
        "pipeline.html",
        form_values=_pipeline_form_values(),
        job=job,
        pipeline_busy=_pipeline_is_busy(),
        messages=flash_messages(),
    )


@app.route("/pipeline/run", methods=["POST"])
@login_required
def run_pipeline():
    values = _pipeline_form_values(request.form)
    command, errors = _pipeline_command(values)
    if errors:
        for error in errors:
            flash(error, "error")
        return (
            render_template(
                "pipeline.html",
                form_values=values,
                job=None,
                pipeline_busy=_pipeline_is_busy(),
                messages=flash_messages(),
            ),
            400,
        )

    try:
        job = _start_pipeline_job(command, str(current_user.id))
    except RuntimeError as exc:
        flash(str(exc), "warning")
        return (
            render_template(
                "pipeline.html",
                form_values=values,
                job=None,
                pipeline_busy=True,
                messages=flash_messages(),
            ),
            409,
        )
    except OSError as exc:
        app.logger.exception("Could not start the Label-Check pipeline")
        flash(f"The pipeline process could not be started: {exc}", "error")
        return (
            render_template(
                "pipeline.html",
                form_values=values,
                job=None,
                pipeline_busy=False,
                messages=flash_messages(),
            ),
            500,
        )

    session["pipeline_job_id"] = job.id
    return redirect(url_for("pipeline_launcher"))


@app.route("/pipeline/jobs/<job_id>/output", methods=["GET"])
@login_required
def pipeline_job_output(job_id: str):
    job = _pipeline_job_for_user(job_id)
    if job is None:
        return jsonify({"error": "Pipeline job not found."}), 404
    try:
        offset = max(0, int(request.args.get("offset", "0")))
    except ValueError:
        return jsonify({"error": "Output offset must be a whole number."}), 400

    with _pipeline_jobs_lock:
        offset = min(offset, len(job.output))
        output = job.output[offset:]
        next_offset = len(job.output)
        status = job.status
        return_code = job.return_code
    return jsonify(
        {
            "output": output,
            "next_offset": next_offset,
            "status": status,
            "return_code": return_code,
        }
    )


@app.route("/inventories", methods=["GET"])
@login_required
def inventories():
    inventory_directory = Path(Config.SCANNER_INVENTORIES)
    inventory_files: List[Path] = []
    directory_error = None

    try:
        if not inventory_directory.is_dir():
            directory_error = (
                f"The scanner inventory directory is unavailable: {inventory_directory}"
            )
        else:
            inventory_files = sorted(
                (
                    path
                    for path in inventory_directory.iterdir()
                    if path.is_file()
                    and not path.is_symlink()
                    and path.suffix.lower() == ".csv"
                ),
                key=lambda path: path.name.casefold(),
            )
    except OSError as exc:
        directory_error = f"The scanner inventory directory could not be read: {exc}"

    selected_name = request.args.get("file", "").strip()
    selected_path = None
    headers: List[str] = []
    rows: List[List[str]] = []
    total_rows = 0
    current_page = 1
    total_pages = 1
    inventory_error = None

    if selected_name:
        files_by_name = {path.name: path for path in inventory_files}
        selected_path = files_by_name.get(selected_name)
        if selected_path is None:
            flash("The selected scanner inventory is unavailable.", "warning")
            return redirect(url_for("inventories"))

        try:
            requested_page = max(1, int(request.args.get("page", "1")))
        except (TypeError, ValueError):
            requested_page = 1

        try:
            (
                headers,
                rows,
                total_rows,
                current_page,
                total_pages,
            ) = _read_inventory_page(selected_path, requested_page)
        except InventoryReadError as exc:
            inventory_error = str(exc)

    return render_template(
        "inventories.html",
        inventory_files=inventory_files,
        selected_name=selected_path.name if selected_path else None,
        headers=headers,
        rows=rows,
        total_rows=total_rows,
        current_page=current_page,
        total_pages=total_pages,
        directory_error=directory_error,
        inventory_error=inventory_error,
        messages=flash_messages(),
    )


@app.route("/sdl", methods=["GET", "POST"])
@login_required
def sdl():
    if request.method == "GET":
        requested_row = request.args.get("edit_row", "").strip()
        if not requested_row:
            return _render_sdl_page()
        try:
            edit_row = int(requested_row)
        except ValueError:
            flash("Invalid SDL row selected.", "error")
            return redirect(url_for("sdl"))
        return _render_sdl_page(edit_row=edit_row)

    submitted_values = _submitted_sdl_form()
    action = request.form.get("action", "add")
    try:
        normalized_values = _validate_sdl_form(submitted_values)
    except SDLValidationError as exc:
        flash(str(exc), "error")
        edit_row = None
        if action == "update":
            try:
                edit_row = int(request.form.get("worksheet_row", ""))
            except ValueError:
                pass
        return _render_sdl_page(
            form_values=submitted_values,
            edit_row=edit_row,
            edit_signature=request.form.get("row_signature", ""),
        )

    workbook = None
    try:
        with _sdl_workbook_lock:
            workbook, worksheet, initialized_headers = _load_sdl_workbook()
            if action == "update":
                try:
                    row_number = int(request.form.get("worksheet_row", ""))
                except ValueError as exc:
                    raise SDLValidationError("Invalid SDL row selected.") from exc
                if row_number < 2 or row_number > worksheet.max_row:
                    raise SDLValidationError("The selected SDL row no longer exists.")
                if all(
                    worksheet.cell(row=row_number, column=column).value is None
                    for column in range(1, len(SDL_HEADERS) + 1)
                ):
                    raise SDLValidationError("The selected SDL row no longer exists.")
                expected_signature = request.form.get("row_signature", "")
                if not expected_signature or expected_signature != _sdl_row_signature(
                    worksheet, row_number
                ):
                    raise SDLValidationError(
                        "This SDL row changed after it was opened. Reload it before saving."
                    )
            elif action == "add":
                row_number = worksheet.max_row + 1
            else:
                raise SDLValidationError("Invalid SDL action.")

            for column, header in enumerate(SDL_HEADERS, start=1):
                if header in normalized_values:
                    worksheet.cell(row=row_number, column=column).value = normalized_values[header]
                elif action == "add" and header in SDL_STATUS_HEADERS:
                    worksheet.cell(row=row_number, column=column).value = False

            for header in ("Date Loaded", "Date Unloaded"):
                worksheet.cell(row=row_number, column=SDL_HEADERS.index(header) + 1).number_format = "yyyy-mm-dd"
            for header in ("Time Loaded", "Time Unloaded"):
                worksheet.cell(row=row_number, column=SDL_HEADERS.index(header) + 1).number_format = "hh:mm"

            _save_sdl_workbook(workbook)
    except (SDLWorkbookError, SDLValidationError) as exc:
        flash(str(exc), "error")
        app.logger.warning("SDL save rejected: %s", exc)
        return _render_sdl_page(
            form_values=submitted_values,
            edit_row=(row_number if action == "update" and "row_number" in locals() else None),
            edit_signature=request.form.get("row_signature", ""),
        )
    except Exception as exc:
        app.logger.exception("Unexpected error while saving the SDL workbook")
        flash(f"The Slide Digitization Log could not be saved: {exc}", "error")
        return _render_sdl_page(
            form_values=submitted_values,
            edit_row=None,
        )
    finally:
        if workbook is not None:
            workbook.close()

    flash(
        "Slide Digitization Log row updated successfully."
        if action == "update"
        else "Slide Digitization Log row added successfully.",
        "success",
    )
    return redirect(url_for("sdl"))


@app.route("/qc", methods=["GET"])
@login_required
def qc():
    context, available_batches, discovery_warnings = _selected_batch()
    if context is None:
        return render_template(
            "batches.html",
            batches=available_batches,
            discovery_warnings=discovery_warnings,
            messages=flash_messages(),
        )

    data_manager = context.data_manager
    queue_manager = context.queue_manager

    _release_expired_leases(context)
    queue_manager.load()

    item_to_display = None
    requested_index_str = request.args.get("index")

    # 1. User requested specific index
    if requested_index_str:
        try:
            idx = int(requested_index_str)
            if 0 <= idx < len(data_manager.data):
                # Release existing leases for this user that are not the requested one
                existing_leases = [
                    l for l in queue_manager.get_all() 
                    if l.leased_by_id == current_user.id and l.status == "leased"
                ]
                for lease in existing_leases:
                    if lease.original_index != idx:
                        lease.status = "pending"
                        lease.leased_by_id = None
                        lease.leased_at = None
                
                qi = queue_manager.get(idx)
                if not qi: 
                    # Should exist if created in init, but if not create ephemeral or fail?
                    # We assume queue is sync'd. 
                    qi = QueueItem(original_index=idx)
                    queue_manager.add(qi)

                if qi.status == "leased" and qi.leased_by_id != current_user.id:
                    flash("This item is currently leased by another user. Viewing in read-only mode.", "warning")
                elif qi.status != "completed":
                    # Acquiring lease
                    qi.status = "leased"
                    qi.leased_by_id = current_user.id
                    qi.leased_at = datetime.datetime.utcnow()
                
                queue_manager.save()
                item_to_display = qi
        except (ValueError, TypeError):
            flash("Invalid index provided in URL.", "error")

    # 2. Check active lease
    if not item_to_display:
        active_lease = next(
            (i for i in queue_manager.get_all() if i.leased_by_id == current_user.id and i.status == "leased"),
            None
        )

        if active_lease:
            item_to_display = active_lease
        else:
            # 3. Get next pending
            # Sort by original_index 
            pending_items = sorted(
                [i for i in queue_manager.get_all() if i.status == "pending"],
                key=lambda x: x.original_index
            )
            
            if pending_items:
                next_pending_item = pending_items[0]
                next_pending_item.status = "leased"
                next_pending_item.leased_by_id = current_user.id
                next_pending_item.leased_at = datetime.datetime.utcnow()
                queue_manager.save()
                item_to_display = next_pending_item
            else:
                # 4. Every unfinished item is currently leased by another user.
                total = len(queue_manager.items)
                done = len([i for i in queue_manager.get_all() if i.status == "completed"])
                return render_template(
                    "qc.html",
                    no_items_available=True,
                    completed_count=done,
                    total_count=total,
                    messages=flash_messages(),
                    batch_id=context.id,
                    batch_name=context.display_name,
                    discovery_warnings=discovery_warnings,
                )

    current_index = item_to_display.original_index
    row_data = data_manager.get_row(current_index)
    if not row_data:
        flash("Error: Database index mismatch with CSV. Reloading data...", "error")
        context.refresh()
        return redirect(url_for("qc", batch=context.id))

    display_row_data = row_data.copy()

    # Pre-fill logic safe lookup
    identifier = display_row_data.get("_identifier")
    if not display_row_data.get("AccessionID") and identifier:
        for r in data_manager.data:
            if r.get("_identifier") == identifier and r.get("AccessionID"):
                display_row_data["AccessionID"] = r["AccessionID"]
                flash(f"Auto-filled Accession ID '{r['AccessionID']}' from a related file.", "info")
                if not display_row_data.get("Stain") and r.get("Stain"):
                    display_row_data["Stain"] = r["Stain"]
                break

    # Image Paths
    label_image_url, macro_image_url = None, None
    label_image_exists, macro_image_exists = False, False

    def resolve_image_path(csv_path_key):
        csv_path = display_row_data.get(csv_path_key)
        if csv_path:
            full_path = data_manager.get_absolute_path(csv_path)
            if full_path and os.path.exists(full_path):
                relative_path = os.path.relpath(full_path, context.root).replace(os.sep, "/")
                return url_for("serve_relative_image", batch=context.id, filepath=relative_path), True
        return None, False

    label_image_url, label_image_exists = resolve_image_path("_label_path")
    macro_image_url, macro_image_exists = resolve_image_path("_macro_path")

    queue_stats = {
        "pending": len([i for i in queue_manager.get_all() if i.status == "pending"]),
        "leased": len([i for i in queue_manager.get_all() if i.status == "leased"]),
        "completed": len([i for i in queue_manager.get_all() if i.status == "completed"]),
    }
    
    recently_completed_items = sorted(
        [i for i in queue_manager.get_all() if i.completed_by_id == current_user.id],
        key=lambda x: x.completed_at if x.completed_at else datetime.datetime.min,
        reverse=True
    )[:5]
    
    # Enrich for template (needs accession_id)
    recently_completed = []
    for r in recently_completed_items:
        rr = data_manager.get_row(r.original_index)
        # Create a proxy object or dict for template
        r_dict = r.to_dict()
        r_dict['accession_id'] = rr.get("AccessionID", "N/A") if rr else "N/A"
        # Overwrite string date with object for strftime support in template
        r_dict['completed_at'] = r.completed_at 
        recently_completed.append(r_dict)

    return render_template(
        "qc.html",
        row=display_row_data,
        original_index=current_index,
        total_original_rows=len(data_manager.data),
        label_img_path=label_image_url,
        macro_img_path=macro_image_url,
        label_img_exists=label_image_exists,
        macro_img_exists=macro_image_exists,
        messages=flash_messages(),
        data_loaded=True,
        queue_stats=queue_stats,
        lease_info=item_to_display,
        datetime=datetime.datetime,
        timedelta=datetime.timedelta,
        recently_completed=recently_completed,
        batch_id=context.id,
        batch_name=context.display_name,
        discovery_warnings=discovery_warnings,
    )


@app.route("/update", methods=["POST"])
@login_required
def update():
    """Handles the form submission for saving corrections."""
    context, _, _ = _selected_batch(allow_completed=True)
    if context is None:
        flash("The selected batch is no longer available.", "warning")
        return redirect(url_for("qc"))
    data_manager = context.data_manager
    queue_manager = context.queue_manager
    if not data_manager.data:
        return redirect(url_for("qc"))
        
    try:
        idx = int(request.form.get("original_index", -1))
        if idx < 0:
            raise ValueError("Invalid index")

        qi = queue_manager.get(idx)

        if not qi:
            flash("Error: Item not found in queue.", "error")
            return redirect(url_for("qc"))

        # --- SAFETY CHECK: LEASE VALIDATION ---
        is_forced_save = False
        
        # Case A: Item is completed.
        if qi.status == "completed":
            flash("Cannot save changes: This item has already been completed.", "error")
            return redirect(url_for("qc"))

        # Case B: I hold the lease.
        if qi.leased_by_id == current_user.id:
            pass # Valid save

        # Case C: Leased by SOMEONE ELSE.
        elif qi.status == "leased" and qi.leased_by_id != current_user.id:
            # Check for lease expiry just in case
            _release_expired_leases(context)
            # Reload queue just to be sure
            qi = queue_manager.get(idx)
            if qi.status == "leased" and qi.leased_by_id != current_user.id:
                flash("SAVE BLOCKED: This item is currently currently leased by another user.", "error")
                return redirect(url_for("qc"))
            # If after refresh it's effectively pending, we fall through to Case D.
            is_forced_save = True

        # Case D: Item is pending (lease expired or never leased).
        elif qi.status == "pending":
            is_forced_save = True # Allowed to pick up

        # --- Update Data ---
        new_values = {
            "AccessionID": request.form.get("accession_id", "").strip(),
            "Stain": request.form.get("stain", "").strip(),
            "BlockNumber": request.form.get("block_number", "").strip(),
            "_is_complete": request.form.get("complete") == "on"
        }
        
        # Validation for completion
        if new_values["_is_complete"]:
            validation_errors = _qc_row_validation_errors(new_values)
            if validation_errors:
                flash(
                    "Cannot mark as complete: " + "; ".join(validation_errors) + ".",
                    "warning",
                )
                new_values["_is_complete"] = False

        # Apply updates
        has_changed = data_manager.update_row(idx, new_values)

        if has_changed:
            current_user.correction_count += 1
            user_manager.save()
            
            if request.form.get("action") == "next" and new_values["_is_complete"]:
                qi.status = "completed"
                qi.completed_by_id = current_user.id
                qi.completed_at = datetime.datetime.utcnow()
            elif is_forced_save:
                qi.status = "leased"
                qi.leased_by_id = current_user.id
                qi.leased_at = datetime.datetime.utcnow()
            
            queue_manager.save()

            try:
                _create_backup(context)
                data_manager.save_data(context.csv_path)
                context.csv_mod_time = context.csv_path.stat().st_mtime

                # --- CHECK IF LIST IS DONE ---
                remaining = len([i for i in queue_manager.get_all() if i.status != "completed"])

                if remaining == 0:
                    invalid_indices = _requeue_invalid_qc_rows(context)
                    if invalid_indices:
                        data_manager.save_data(context.csv_path)
                        context.csv_mod_time = context.csv_path.stat().st_mtime
                        app.logger.warning(
                            "Final QC validation returned %d row(s) to the queue for batch %s.",
                            len(invalid_indices),
                            context.id,
                        )
                        flash(
                            f"Final validation returned {len(invalid_indices)} slide(s) "
                            "to the QC queue because required values were missing or invalid.",
                            "warning",
                        )
                    else:
                        app.logger.info("All items passed final validation. Creating final backup.")
                        _create_backup(context, suffix="FINAL_COMPLETED")
                        try:
                            context.mark_qc_complete()
                        except DataSaveError as exc:
                            app.logger.error("QC status update failed: %s", exc)
                            flash(
                                "Slide changes were saved, but the batch could not be marked as QC complete.",
                                "error",
                            )
                            return redirect(url_for("qc", batch=context.id))
                        _start_renaming_job(context)
                        flash("🎉 ALL ITEMS COMPLETED! A final comprehensive backup has been created.", "success")
                else:
                    flash("Changes saved successfully.", "success")

            except Exception as e:
                app.logger.error(f"Save operation failed: {e}")
                flash("CRITICAL: Error saving changes to the CSV file.", "error")

        return redirect(url_for("qc", batch=context.id))

    except Exception as e:
        app.logger.error(f"Update failed: {e}")
        flash("An error occurred during the update.", "error")
        return redirect(url_for("qc"))


@app.route("/history")
@login_required
def history():
    context, _, _ = _selected_batch(allow_completed=True)
    if context is None:
        flash("Choose a batch to view its history.", "warning")
        return redirect(url_for("qc"))
    data_manager = context.data_manager
    queue_manager = context.queue_manager
    history_items = sorted(
        [i for i in queue_manager.get_all() if i.completed_by_id == current_user.id],
        key=lambda x: x.completed_at if x.completed_at else datetime.datetime.min,
        reverse=True
    )
    
    # Enhance for template
    display_history = []
    for item in history_items:
        d = item.to_dict()
        row = data_manager.get_row(item.original_index)
        d['accession_id'] = row.get("AccessionID", "N/A") if row else "N/A"
        d['completed_at'] = item.completed_at
        display_history.append(d)

    return render_template(
        "history.html", completed_items=display_history, messages=flash_messages(),
        batch_id=context.id, batch_name=context.display_name,
    )


@app.route("/release", methods=["POST"])
@login_required
def release_lease():
    context, _, _ = _selected_batch(allow_completed=True)
    if context is None:
        flash("The selected batch is no longer available.", "warning")
        return redirect(url_for("qc"))
    queue_manager = context.queue_manager
    leases = [
        l for l in queue_manager.get_all() 
        if l.leased_by_id == current_user.id and l.status == "leased"
    ]
    
    if leases:
        for lease in leases:
            lease.status = "pending"
            lease.leased_by_id = None
            lease.leased_at = None
        queue_manager.save()
        flash(f"Successfully released {len(leases)} item(s) back to the queue.", "info")
        
    return redirect(url_for("qc", batch=context.id))


@app.route("/search", methods=["POST"])
@login_required
def search():
    context, _, _ = _selected_batch()
    if context is None:
        return redirect(url_for("qc"))
    data_manager = context.data_manager
    if not data_manager.data:
        return redirect(url_for("qc"))
        
    search_term = request.form.get("search_term", "").strip().lower()
    if not search_term:
        return redirect(url_for("qc"))

    for i, row in enumerate(data_manager.data):
        if (
            search_term in row.get("AccessionID", "").lower() or
            search_term in row.get("_identifier", "").lower() or
            search_term == row.get("BlockNumber", "").lower()
        ):
            return redirect(url_for("qc", batch=context.id, index=i))

    flash(f"No item found matching '{search_term}'.", "warning")
    return redirect(url_for("qc", batch=context.id))


@app.route("/data_images/<batch>/<path:filepath>")
@login_required
def serve_relative_image(batch: str, filepath: str):
    batches, _ = discover_batches()
    context = next((item for item in batches if item.id == batch), None)
    if context is None:
        return "Batch not found.", 404
    abs_file_path = context.data_manager.get_absolute_path(filepath)
    if not abs_file_path:
        app.logger.warning("Blocked invalid image path for batch %s: %s", batch, filepath)
        return "Access denied: Invalid file path.", 403
    if not os.path.isfile(abs_file_path):
        return "Image not found on server.", 404

    directory, filename = os.path.split(abs_file_path)
    return send_from_directory(directory, filename)


# ==============================================================================
# 11. VERSIONED PIPELINE API
# ==============================================================================
API_PIPELINE_FIELDS = {
    "input_dir", "output_dir", "start_from", "end_at", "input_mode",
    "macro_workers", "macro_extensions", "macro_image_extensions",
    "thumbnail_width", "thumbnail_height", "ocr_workers", "ocr_use_cpu",
    "naming_accession_pattern", "naming_workers",
}


def _api_pipeline_values(payload: Any) -> Tuple[Optional[Dict[str, str]], List[str]]:
    if not isinstance(payload, dict):
        return None, ["The request body must be a JSON object."]
    errors = []
    unknown = sorted(set(payload) - API_PIPELINE_FIELDS)
    if unknown:
        errors.append(f"Unknown fields: {', '.join(unknown)}.")
    for required in ("input_dir", "output_dir"):
        if required not in payload:
            errors.append(f"{required} is required.")

    integer_fields = {
        "start_from", "end_at", "macro_workers", "thumbnail_width",
        "thumbnail_height", "ocr_workers", "naming_workers",
    }
    string_fields = {"input_dir", "output_dir", "input_mode", "naming_accession_pattern"}
    extension_fields = {"macro_extensions", "macro_image_extensions"}
    for field in integer_fields & payload.keys():
        if isinstance(payload[field], bool) or not isinstance(payload[field], int):
            errors.append(f"{field} must be an integer.")
    for field in string_fields & payload.keys():
        if not isinstance(payload[field], str):
            errors.append(f"{field} must be a string.")
    for field in extension_fields & payload.keys():
        value = payload[field]
        if not isinstance(value, list) or not value or any(not isinstance(item, str) for item in value):
            errors.append(f"{field} must be a non-empty array of strings.")
    if "ocr_use_cpu" in payload and not isinstance(payload["ocr_use_cpu"], bool):
        errors.append("ocr_use_cpu must be a boolean.")
    if errors:
        return None, errors

    values = dict(PIPELINE_FORM_DEFAULTS)
    for key, value in payload.items():
        if key in integer_fields:
            values[key] = str(value)
        elif key in extension_fields:
            values[key] = ", ".join(value)
        elif key == "ocr_use_cpu":
            values[key] = "on" if value else ""
        else:
            values[key] = value
    return values, []


@app.route("/api/v1/pipeline/jobs", methods=["POST"])
@_require_api_scope("pipeline:run", "submit")
def api_create_pipeline_job():
    if not request.is_json:
        return _api_problem(415, "unsupported_media_type", "JSON required", "Use Content-Type: application/json.")
    idempotency_key = request.headers.get("Idempotency-Key", "")
    if not re.fullmatch(r"[\x21-\x7E]{1,128}", idempotency_key):
        return _api_problem(422, "invalid_idempotency_key", "Invalid idempotency key", "Idempotency-Key must contain 1–128 printable non-space ASCII characters.")
    payload = request.get_json(silent=True)
    values, shape_errors = _api_pipeline_values(payload)
    if shape_errors:
        return _api_problem(422, "validation_error", "Invalid pipeline request", " ".join(shape_errors))
    assert values is not None
    payload_hash = hashlib.sha256(
        json.dumps(values, sort_keys=True, separators=(",", ":")).encode("utf-8")
    ).hexdigest()
    existing = api_store.find_idempotent(g.api_token["token_id"], idempotency_key)
    if existing is not None:
        if existing["payload_hash"] != payload_hash:
            return _api_problem(409, "idempotency_conflict", "Idempotency conflict", "This key was already used with a different request.")
        response = jsonify({"data": _api_job_document(dict(existing))})
        response.status_code = 202
        response.headers["Location"] = url_for("api_pipeline_job", job_id=existing["job_id"], _external=True)
        response.headers["Idempotency-Replayed"] = "true"
        return response

    command, validation_errors = _pipeline_command(values)
    if validation_errors:
        return _api_problem(422, "validation_error", "Invalid pipeline request", " ".join(validation_errors))
    assert command is not None
    job_id = str(uuid.uuid4())
    try:
        job = _start_pipeline_job(
            command,
            str(g.api_user.id),
            job_id=job_id,
            request_values=values,
            token_id=g.api_token["token_id"],
            idempotency_key=idempotency_key,
            payload_hash=payload_hash,
        )
    except RuntimeError as exc:
        raced_job = api_store.find_idempotent(g.api_token["token_id"], idempotency_key)
        if raced_job is not None and raced_job["payload_hash"] == payload_hash:
            response = jsonify({"data": _api_job_document(dict(raced_job))})
            response.status_code = 202
            response.headers["Location"] = url_for(
                "api_pipeline_job", job_id=raced_job["job_id"], _external=True
            )
            response.headers["Idempotency-Replayed"] = "true"
            return response
        return _api_problem(409, "pipeline_busy", "Pipeline busy", str(exc))
    except OSError:
        app.logger.exception("API could not start the Label-Check pipeline")
        return _api_problem(500, "pipeline_launch_failed", "Pipeline launch failed", "The pipeline process could not be started.")
    record = api_store.get_job(job.id)
    response = jsonify({"data": _api_job_document(record)})
    response.status_code = 202
    response.headers["Location"] = url_for("api_pipeline_job", job_id=job.id, _external=True)
    return response


def _authorized_api_job(job_id: str) -> Tuple[Optional[Dict[str, Any]], Optional[Any]]:
    record = api_store.get_job(job_id)
    if record is None or (
        record["owner_id"] != str(g.api_user.id) and not g.api_user.is_admin
    ):
        return None, _api_problem(404, "job_not_found", "Job not found", "The requested pipeline job was not found.")
    return record, None


@app.route("/api/v1/pipeline/jobs/<job_id>", methods=["GET"])
@_require_api_scope("pipeline:read")
def api_pipeline_job(job_id: str):
    record, error = _authorized_api_job(job_id)
    if error is not None:
        return error
    return jsonify({"data": _api_job_document(record)})


@app.route("/api/v1/pipeline/jobs/<job_id>/output", methods=["GET"])
@_require_api_scope("pipeline:read")
def api_pipeline_job_output(job_id: str):
    record, error = _authorized_api_job(job_id)
    if error is not None:
        return error
    try:
        offset = int(request.args.get("offset", "0"))
        limit = int(request.args.get("limit", str(app.config["API_OUTPUT_DEFAULT_LIMIT"])))
    except ValueError:
        return _api_problem(422, "invalid_pagination", "Invalid pagination", "offset and limit must be integers.")
    if offset < 0 or limit < 1 or limit > app.config["API_OUTPUT_MAX_LIMIT"]:
        return _api_problem(422, "invalid_pagination", "Invalid pagination", f"offset must be non-negative and limit must be 1–{app.config['API_OUTPUT_MAX_LIMIT']}.")
    output_path = record["output_path"]
    try:
        size = os.path.getsize(output_path)
        offset = min(offset, size)
        with open(output_path, "rb") as output_file:
            output_file.seek(offset)
            raw = output_file.read(limit)
    except OSError:
        return _api_problem(500, "output_unavailable", "Output unavailable", "The pipeline output could not be read.")
    while raw:
        try:
            output = raw.decode("utf-8")
            break
        except UnicodeDecodeError as exc:
            if exc.reason == "unexpected end of data":
                raw = raw[:exc.start]
            else:
                output = raw.decode("utf-8", errors="replace")
                break
    else:
        output = ""
    next_offset = offset + len(raw)
    return jsonify(
        {
            "data": {
                "job_id": job_id,
                "output": output,
                "offset": offset,
                "next_offset": next_offset,
                "eof": next_offset >= size,
                "status": record["status"],
            }
        }
    )


@app.route("/api/v1/openapi.json", methods=["GET"])
@_require_api_scope("pipeline:read")
def api_openapi_document():
    contract_path = Path(__file__).resolve().with_name("openapi.json")
    try:
        with contract_path.open("r", encoding="utf-8") as contract_file:
            return jsonify(json.load(contract_file))
    except (OSError, json.JSONDecodeError):
        app.logger.exception("OpenAPI contract is unavailable")
        return _api_problem(500, "contract_unavailable", "Contract unavailable", "The OpenAPI contract could not be loaded.")


# ==============================================================================
# 12. CLI COMMANDS
# ==============================================================================
@app.cli.group("api-token")
def api_token_cli():
    """Manage scoped personal access tokens for the pipeline API."""


@api_token_cli.command("create")
@click.argument("user_id")
@click.option("--label", required=True, help="Human-readable credential label.")
@click.option(
    "--scope",
    "scopes",
    multiple=True,
    type=click.Choice(["pipeline:read", "pipeline:run"]),
    default=("pipeline:read", "pipeline:run"),
    show_default=True,
)
@click.option("--expires-days", type=click.IntRange(min=1), default=90, show_default=True)
def create_api_token(user_id: str, label: str, scopes: Tuple[str, ...], expires_days: int):
    if user_manager.get(user_id) is None:
        raise click.ClickException(f"Unknown user: {user_id}")
    raw_token, record = api_store.create_token(user_id, label, list(scopes), expires_days)
    click.echo(f"Token ID: {record['token_id']}")
    click.echo(f"Expires: {record['expires_at']}")
    click.echo("Token (shown once):")
    click.echo(raw_token)


@api_token_cli.command("list")
@click.option("--user", "user_id")
def list_api_tokens(user_id: Optional[str]):
    records = api_store.list_tokens(user_id)
    if not records:
        click.echo("No API tokens found.")
        return
    for record in records:
        state = "revoked" if record["revoked_at"] else "active"
        click.echo(
            f"{record['token_id']}\t{record['user_id']}\t{record['label']}\t"
            f"{','.join(record['scopes'])}\t{record['expires_at']}\t{state}"
        )


@api_token_cli.command("revoke")
@click.argument("token_id")
def revoke_api_token(token_id: str):
    if not api_store.revoke_token(token_id):
        raise click.ClickException("Active token not found.")
    click.echo(f"Revoked token {token_id}.")


@api_token_cli.command("rotate")
@click.argument("token_id")
@click.option("--expires-days", type=click.IntRange(min=1), default=90, show_default=True)
def rotate_api_token(token_id: str, expires_days: int):
    record = next((item for item in api_store.list_tokens() if item["token_id"] == token_id), None)
    if record is None or record["revoked_at"]:
        raise click.ClickException("Active token not found.")
    raw_token, replacement = api_store.create_token(
        record["user_id"], record["label"], record["scopes"], expires_days
    )
    api_store.revoke_token(token_id)
    click.echo(f"Revoked token {token_id}; replacement ID: {replacement['token_id']}")
    click.echo("Token (shown once):")
    click.echo(raw_token)


@app.cli.command("init-db")
@with_appcontext
def init_db_command():
    print("--- Initializing App Persistence (CSV) ---")
    
    # Init Users
    if not user_manager.get("admin"):
        u = User(id="admin", password_hash="", is_admin=True)
        u.set_password(Config.ADMIN_DEFAULT_PASSWORD)
        user_manager.add(u)
        print(f"Created default 'admin' user in {Config.USERS_CSV_PATH}")
    else:
        print("'admin' user already exists.")

    batches, warnings = discover_batches()
    print(f"Discovered and initialized {len(batches)} valid batch(es).")
    for warning in warnings:
        print(f"WARNING: {warning}")

    print("--- Initialization complete. ---")


if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0")
